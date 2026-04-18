import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent / "core"))
import os
import logging
import threading
import tempfile
from datetime import datetime, date, timedelta
from fastapi import APIRouter, Depends, HTTPException
logger = logging.getLogger(__name__)
from pydantic import BaseModel
from supabase import create_client
from backend.auth import get_current_user
from backend.jobs import create_job, update_job, get_job, get_user_jobs
from dotenv import load_dotenv
load_dotenv()

router = APIRouter(prefix="/api/reports", tags=["reports"])
SUPABASE_URL         = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_KEY = os.getenv("SUPABASE_SERVICE_KEY")

def get_supabase():
    return create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)


def _parse_report_dates(start_date: str, end_date: str) -> tuple[date, date]:
    """Validate start_date / end_date strings and return parsed date objects.

    Raises HTTPException(400) on bad format, reversed range, or out-of-band
    values (>20y in the past or >1y in the future).
    """
    try:
        s = datetime.strptime(start_date, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="Invalid date format, expected YYYY-MM-DD")
    try:
        e = datetime.strptime(end_date, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="Invalid date format, expected YYYY-MM-DD")

    if s > e:
        raise HTTPException(status_code=400, detail="start_date must be on or before end_date")

    today    = date.today()
    earliest = today - timedelta(days=365 * 20)
    latest   = today + timedelta(days=365)
    if s < earliest or e > latest:
        raise HTTPException(
            status_code=400,
            detail="Dates must be within 20 years past and 1 year future",
        )
    return s, e


class GenerateRequest(BaseModel):
    realm_id:          str
    start_date:        str
    end_date:          str
    dimension:         str = "none"
    selected_maps:      list[str] = []
    include_gl_detail:  bool = False
    include_portal_data: bool = False
    include_ar_aging:   bool = False
    include_ap_aging:   bool = False

def run_report_job(job_id: str, user_id: str, realm_id: str,
                   start_date: str, end_date: str, dimension: str,
                   selected_maps: list[str] | None = None,
                   include_gl_detail: bool = False,
                   include_portal_data: bool = False,
                   include_ar_aging: bool = False,
                   include_ap_aging: bool = False):
    """Run in a background thread — fetches QBO data and generates Excel file."""
    try:
        update_job(job_id, status="running")
        import sys
        sys.path.insert(0, str(Path(__file__).parent / "core"))
        from gl_extractor import generate_lite
        supabase = get_supabase()
        token_result = supabase.table("qbo_tokens").select(
            "access_token, refresh_token, company_name"
        ).eq("user_id", user_id).eq("realm_id", realm_id).execute()
        if not token_result.data:
            update_job(job_id, status="failed", error="No QBO connection found")
            return
        tokens        = token_result.data[0]
        company_name  = tokens.get("company_name", "") or realm_id
        access_token  = tokens["access_token"]
        refresh_token = tokens["refresh_token"]
        logger.info(f"company_name from tokens: '{company_name}'")

        from datetime import datetime, timedelta
        import re as _re

        # Always refresh QBO token before running
        try:
            import base64, httpx as _httpx
            client_id     = os.getenv("QBO_CLIENT_ID", "")
            client_secret = os.getenv("QBO_CLIENT_SECRET", "")
            credentials   = base64.b64encode(f"{client_id}:{client_secret}".encode()).decode()
            refresh_resp  = _httpx.post(
                "https://oauth.platform.intuit.com/oauth2/v1/tokens/bearer",
                headers={
                    "Authorization": f"Basic {credentials}",
                    "Content-Type":  "application/x-www-form-urlencoded",
                    "Accept":        "application/json",
                },
                data={"grant_type": "refresh_token", "refresh_token": refresh_token},
                timeout=30,
            )
            if refresh_resp.status_code == 200:
                nt            = refresh_resp.json()
                access_token  = nt["access_token"]
                refresh_token = nt.get("refresh_token", refresh_token)
                new_expiry    = (datetime.utcnow() + timedelta(
                    seconds=nt.get("expires_in", 3600))).isoformat()
                supabase.table("qbo_tokens").update({
                    "access_token":  access_token,
                    "refresh_token": refresh_token,
                    "expires_at":    new_expiry,
                    "updated_at":    datetime.utcnow().isoformat(),
                }).eq("user_id", user_id).eq("realm_id", realm_id).execute()
                logger.info("QBO token refreshed successfully")
            else:
                logger.warning(f"QBO token refresh failed: {refresh_resp.status_code}")
        except Exception as _rfe:
            logger.warning(f"QBO token refresh error: {_rfe}")

        import qbo_client
        qbo_client.set_override_tokens({
            "realm_id":      realm_id,
            "access_token":  access_token,
            "refresh_token": refresh_token,
        })
        qbo_client.get_environment = lambda: "production"
        logger.info(f"Override tokens set for realm_id={realm_id}")

        clean_name = _re.sub(r'[^\w]', '_', company_name).strip('_').upper()
        file_name  = f"{clean_name}_{start_date[:7]}_{end_date[:7]}.xlsx"

        def progress_fn(msg):
            msg = str(msg).strip()
            if msg and not msg.startswith('[progress]'):
                update_job(job_id, progress=msg)

        with tempfile.TemporaryDirectory() as tmpdir:
            result = generate_lite(
                alias=realm_id,
                start_date=start_date,
                end_date=end_date,
                output_mode="new",
                output_folder=tmpdir,
                file_name=file_name,
                dimension=dimension,
                progress_fn=progress_fn,
                include_gl_detail=include_gl_detail,
                include_ar_aging=include_ar_aging,
                include_ap_aging=include_ap_aging,
                company_name=company_name,
            )
            file_path = result["path"]

            # ── Append mapping columns ────────────────────────────────
            if selected_maps:
                try:
                    logger.info(f"Applying maps: {selected_maps}")
                    mapping_result = supabase.table("mappings").select("account_maps").eq(
                        "user_id", user_id
                    ).eq("realm_id", realm_id).execute()
                    account_maps  = (mapping_result.data[0].get("account_maps", [])
                                     if mapping_result.data else [])
                    maps_to_apply = [m for m in account_maps
                                     if m.get("map_name", "") in selected_maps]

                    if maps_to_apply:
                        import openpyxl as _ox
                        from openpyxl.styles import Font, PatternFill, Alignment
                        from openpyxl.utils import get_column_letter
                        from openpyxl.formatting.rule import CellIsRule
                        from collections import defaultdict
                        from datetime import datetime as _dt
                        import calendar as _cal

                        wb = _ox.load_workbook(file_path)

                        if "Normal" in wb.style_names:
                            wb._named_styles["Normal"].font = Font(name="Arial", size=10)

                        for sn in wb.sheetnames:
                            wb[sn].sheet_view.showGridLines = False

                        # Shared styles
                        HDR_FONT    = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                        HDR_FILL    = PatternFill("solid", fgColor="337E8D")
                        GL_HDR_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                        GL_HDR_FILL = PatternFill("solid", fgColor="337E8D")
                        BOLD        = Font(name="Arial", size=10, bold=True)
                        BOLD_LG     = Font(name="Arial", size=10, bold=True)
                        SEC_FILL    = PatternFill("solid", fgColor="D9E1F2")
                        SUBTOT_FILL = PatternFill("solid", fgColor="EEF2F7")
                        NUM_FMT     = '#,##0.00_);(#,##0.00);"-"??;@'
                        RED_FONT    = Font(name="Arial", size=10, bold=True)
                        RED_FILL    = PatternFill("solid", fgColor="FFC7CE")
                        GRN_FONT    = Font(name="Arial", size=10, bold=True)
                        GRN_FILL    = PatternFill("solid", fgColor="C6EFCE")
                        LINK_FONT   = Font(name="Arial", size=10, color="276221")

                        def build_lookup(m):
                            out = {}
                            for grp in m.get("groups", []):
                                gname = grp.get("group_name", "")
                                sect  = grp.get("pl_section") or grp.get("bs_section") or ""
                                for a in grp.get("accounts", []):
                                    name = (a.get("account_name", "") if isinstance(a, dict)
                                            else str(a)).strip()
                                    if name:
                                        out[name] = (gname, sect)
                            return out

                        _GL_HDR_R = 5   # column headers now at row 5
                        _GL_DATA_R = 6  # data starts at row 6
                        _DC = 2         # data columns start at B (buffer col A)

                        # ── Append to IS GL Detail, BS GL Detail, BS Balances ──
                        for tab in ("IS GL Summary", "BS GL Summary", "IS GL Detail", "BS GL Detail", "BS Balances"):
                            if tab not in wb.sheetnames:
                                continue
                            ws  = wb[tab]
                            hdr = [ws.cell(_GL_HDR_R, c).value for c in range(1, ws.max_column + 1)]
                            acct_col = None
                            for candidate in ("Account Name", "Account"):
                                if candidate in hdr:
                                    acct_col = hdr.index(candidate) + 1
                                    break
                            if not acct_col:
                                continue
                            base = ws.max_column + 1
                            for mi, m in enumerate(maps_to_apply):
                                lkp   = build_lookup(m)
                                gc    = base + mi * 2
                                sc    = base + mi * 2 + 1
                                mname = m.get("map_name", "")
                                ws.cell(_GL_HDR_R, gc, f"{mname} - Account Group").font = GL_HDR_FONT
                                ws.cell(_GL_HDR_R, gc).fill = GL_HDR_FILL
                                ws.cell(_GL_HDR_R, sc, f"{mname} - Statement Section").font = GL_HDR_FONT
                                ws.cell(_GL_HDR_R, sc).fill = GL_HDR_FILL
                                for ri in range(_GL_DATA_R, ws.max_row + 1):
                                    v = ws.cell(ri, acct_col).value
                                    if not v:
                                        continue
                                    match = lkp.get(str(v).strip())
                                    if match:
                                        ws.cell(ri, gc, match[0])
                                        ws.cell(ri, sc, match[1])

                            # Autofit mapping columns after all data is written
                            for ci in range(base, ws.max_column + 1):
                                max_len = 0
                                for ri in range(_GL_HDR_R, ws.max_row + 1):
                                    cv = ws.cell(ri, ci).value
                                    if cv is not None:
                                        max_len = max(max_len, len(str(cv)))
                                ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 60)

                        # ── Mapping Reference tab ──────────────────────────
                        try:
                            if "Mapping Reference" in wb.sheetnames:
                                del wb["Mapping Reference"]
                            ws_ref = wb.create_sheet("Mapping Reference")
                            ws_ref.sheet_view.showGridLines = False
                            ws_ref.column_dimensions["A"].width = 0.63
                            ws_ref.cell(1, 1, company_name).font = Font(name="Arial", size=10, bold=True)
                            ws_ref.cell(2, 1, "Mapping Reference").font = Font(name="Arial", size=10)

                            # Fetch full COA from QBO (all accounts, not just those with activity)
                            from qbo_client import fetch_accounts as _fetch_coa
                            _IS_TYPES = {"Income", "Expense", "Cost of Goods Sold", "Other Income", "Other Expense"}
                            _BS_TYPES = {"Bank", "Accounts Receivable", "Other Current Asset", "Fixed Asset",
                                         "Other Asset", "Accounts Payable", "Credit Card", "Other Current Liability",
                                         "Long Term Liability", "Equity"}
                            _coa_raw = _fetch_coa(realm_id) or []
                            _ref_accounts = {}
                            for _a in _coa_raw:
                                _name = str(_a.get("FullyQualifiedName", _a.get("Name", "")) or "").strip()
                                _atype = str(_a.get("AccountType", "") or "")
                                _anum = str(_a.get("AcctNum", "") or "").strip()
                                _display = f"{_anum} {_name}".strip() if _anum else _name
                                if not _name:
                                    continue
                                if _atype in _IS_TYPES:
                                    _stmt = "IS"
                                elif _atype in _BS_TYPES:
                                    _stmt = "BS"
                                else:
                                    _stmt = ""
                                _ref_accounts[_name] = {"type": _atype, "stmt": _stmt, "display": _display}

                            # Sort: IS first, then BS; within each by type then display name
                            _sorted_accts = sorted(
                                _ref_accounts.items(),
                                key=lambda x: (0 if x[1]["stmt"] == "IS" else 1 if x[1]["stmt"] == "BS" else 2, x[1]["type"], x[1].get("display", x[0]))
                            )

                            # Build lookups for each map
                            _map_lookups = []
                            for m in maps_to_apply:
                                _map_lookups.append((m.get("map_name", ""), build_lookup(m)))

                            # Headers at row 5
                            _ref_headers = ["QBO Account", "Account Type", "Statement"]
                            for mn, _ in _map_lookups:
                                _ref_headers.append(f"{mn} \u2014 Group")
                                _ref_headers.append(f"{mn} \u2014 Section")

                            for ci, h in enumerate(_ref_headers, _DC):
                                c = ws_ref.cell(5, ci, h)
                                c.font = HDR_FONT; c.fill = HDR_FILL
                                c.alignment = Alignment(horizontal="center" if ci > _DC else "left")

                            # Data rows
                            _rr = 6
                            for acct_name, info in _sorted_accts:
                                _disp = info.get("display", acct_name)
                                ws_ref.cell(_rr, _DC, _disp).font = Font(name="Arial", size=10)
                                ws_ref.cell(_rr, _DC + 1, info["type"]).font = Font(name="Arial", size=10)
                                ws_ref.cell(_rr, _DC + 2, info["stmt"]).font = Font(name="Arial", size=10)
                                for mi, (mn, lkp) in enumerate(_map_lookups):
                                    gc = _DC + 3 + mi * 2
                                    sc = _DC + 4 + mi * 2
                                    match = lkp.get(acct_name) or lkp.get(_disp)
                                    if match:
                                        ws_ref.cell(_rr, gc, match[0]).font = Font(name="Arial", size=10)
                                        ws_ref.cell(_rr, sc, match[1]).font = Font(name="Arial", size=10)
                                    else:
                                        ws_ref.cell(_rr, gc, "Unmapped").font = Font(name="Arial", size=10, color="999999")
                                _rr += 1

                            # Autofit columns
                            for ci in range(_DC, _DC + len(_ref_headers)):
                                mx = 0
                                for r in range(5, _rr):
                                    v = ws_ref.cell(r, ci).value
                                    if v is not None:
                                        mx = max(mx, len(str(v)))
                                ws_ref.column_dimensions[get_column_letter(ci)].width = min(max(mx + 4, 15), 40)

                            ws_ref.freeze_panes = "A6"
                        except Exception as _ref_err:
                            logger.warning(f"Mapping Reference tab failed: {_ref_err}")

                        # ── Map Summary tab ────────────────────────────────
                        if "Map Summary" in wb.sheetnames:
                            del wb["Map Summary"]
                        ws_sum = wb.create_sheet("Map Summary")
                        ws_sum.sheet_view.showGridLines = False
                        _first_map = maps_to_apply[0].get("map_name", "") if maps_to_apply else ""
                        ws_sum.cell(1, 1, company_name).font = Font(name="Arial", size=10, bold=True)
                        ws_sum.cell(2, 1, f"{_first_map} \u2014 Summary").font = Font(name="Arial", size=10)
                        cr = 5  # current row (rows 1-4 are global header)

                        for m in maps_to_apply:
                            map_name  = m.get("map_name", "")
                            grp_label = f"{map_name} - Account Group"
                            sec_label = f"{map_name} - Statement Section"

                            # ── IS Section ──────────────────────────────────
                            if "IS GL Summary" not in wb.sheetnames:
                                continue
                            ws_is  = wb["IS GL Summary"]
                            hdr_is = [ws_is.cell(_GL_HDR_R, c).value
                                      for c in range(1, ws_is.max_column + 1)]

                            # Resolve IS GL Summary columns dynamically
                            def _cl(hdrs, name):
                                for i, h in enumerate(hdrs):
                                    if str(h or "").strip() == name:
                                        return (i + 1, get_column_letter(i + 1))
                                return (None, None)

                            is_amt_col_i, is_amt_col_l = _cl(hdr_is, "Amount")
                            is_mon_col_i, is_mon_col_l = _cl(hdr_is, "Month")
                            _, is_sec_col_l = _cl(hdr_is, sec_label)
                            _, is_grp_col_l = _cl(hdr_is, grp_label)

                            if not is_mon_col_i or not is_amt_col_i or not is_sec_col_l:
                                continue

                            # Collect unique months
                            month_keys    = []
                            month_display = {}
                            month_dates   = {}
                            for ri in range(_GL_DATA_R, ws_is.max_row + 1):
                                mv = ws_is.cell(ri, is_mon_col_i).value
                                if mv is None:
                                    continue
                                if hasattr(mv, 'strftime'):
                                    mk = mv.strftime("%Y-%m")
                                    if mk not in month_dates:
                                        month_dates[mk] = mv
                                else:
                                    try:
                                        mk = _dt.strptime(str(mv)[:7], "%Y-%m").strftime("%Y-%m")
                                    except Exception:
                                        mk = str(mv)[:7]
                                if mk not in month_keys:
                                    month_keys.append(mk)
                                if mk not in month_display:
                                    try:
                                        month_display[mk] = _dt.strptime(mk, "%Y-%m").strftime("%b %Y")
                                    except Exception:
                                        month_display[mk] = mk
                            month_keys = sorted(month_keys)
                            num_mo     = len(month_keys)
                            tot_col    = num_mo + _DC + 1  # months start at col 3

                            # IS groups from mapping definition, sorted by section
                            IS_ORDER = ["Revenue", "COS", "Cost of Goods Sold",
                                        "Sales & Marketing", "Operating Expenses",
                                        "Other Income", "Other Expense", "Other"]
                            is_groups = []
                            seen_ig   = set()
                            for grp in m.get("groups", []):
                                gname = grp.get("group_name", "")
                                stmt  = grp.get("statement", "IS")
                                sec   = grp.get("pl_section", "") or grp.get("bs_section", "")
                                if gname and stmt == "IS" and gname not in seen_ig:
                                    seen_ig.add(gname)
                                    is_groups.append((gname, sec))

                            is_groups.sort(key=lambda g: IS_ORDER.index(g[1])
                                           if g[1] in IS_ORDER else 99)

                            # Title
                            ws_sum.cell(cr, _DC,
                                f"{map_name} \u2014 Income Statement").font = BOLD_LG
                            cr += 1

                            # Header — col A blank, then month labels, then Total
                            ws_sum.cell(cr, _DC, "").font = HDR_FONT
                            ws_sum.cell(cr, _DC).fill = HDR_FILL
                            for ci, mk in enumerate(month_keys, _DC + 1):
                                c = ws_sum.cell(cr, ci, month_display.get(mk, mk))
                                c.font = HDR_FONT
                                c.fill = HDR_FILL
                                c.alignment = Alignment(horizontal="center")
                            ws_sum.cell(cr, tot_col, "Total").font = HDR_FONT
                            ws_sum.cell(cr, tot_col).fill = HDR_FILL
                            cr += 1

                            # Aggregate by section — one row per P&L section
                            INCOME_SECS  = {"Revenue", "Other Income"}
                            EXPENSE_SECS = {"COS", "Cost of Goods Sold", "Sales & Marketing", "Operating Expenses", "Other Expense", "Other"}

                            section_order = []
                            seen_secs = set()
                            for gname, sec in is_groups:
                                if sec and sec not in seen_secs:
                                    seen_secs.add(sec)
                                    section_order.append(sec)
                            section_order.sort(key=lambda s: IS_ORDER.index(s) if s in IS_ORDER else 99)

                            data_rows = []  # [(row_number, section_name)]
                            for sec in section_order:
                                ws_sum.cell(cr, _DC, sec)
                                data_rows.append((cr, sec))

                                for ci, mk in enumerate(month_keys, _DC + 1):
                                    mv = month_dates.get(mk)
                                    if mv and hasattr(mv, 'year'):
                                        date_f = f"DATE({mv.year},{mv.month},{mv.day})"
                                    else:
                                        try:
                                            d      = _dt.strptime(mk, "%Y-%m")
                                            last   = _cal.monthrange(d.year, d.month)[1]
                                            date_f = f"DATE({d.year},{d.month},{last})"
                                        except Exception:
                                            date_f = f'"{mk}"'
                                    formula = (
                                        f"=SUMIFS('IS GL Summary'!${is_amt_col_l}:${is_amt_col_l},"
                                        f"'IS GL Summary'!${is_sec_col_l}:${is_sec_col_l},\"{sec}\","
                                        f"'IS GL Summary'!${is_mon_col_l}:${is_mon_col_l},{date_f})"
                                    )
                                    ws_sum.cell(cr, ci, formula).number_format = NUM_FMT

                                sl = get_column_letter(_DC + 1)
                                el = get_column_letter(1 + num_mo)
                                ws_sum.cell(cr, tot_col, f"=SUM({sl}{cr}:{el}{cr})").number_format = NUM_FMT
                                cr += 1

                            # Net Income — income sections minus expense sections
                            ni_row = cr
                            ws_sum.cell(cr, _DC, "Net Income").font = BOLD
                            for ci in range(_DC + 1, tot_col + 1):
                                cl       = get_column_letter(ci)
                                inc_refs = "+".join(f"{cl}{r}" for r, s in data_rows if s in INCOME_SECS)
                                exp_refs = "+".join(f"{cl}{r}" for r, s in data_rows if s in EXPENSE_SECS)
                                if inc_refs and exp_refs:
                                    formula = f"={inc_refs}-({exp_refs})"
                                elif inc_refs:
                                    formula = f"={inc_refs}"
                                else:
                                    formula = "=0"
                                c = ws_sum.cell(cr, ci, formula)
                                c.number_format = NUM_FMT
                                c.font = BOLD
                            cr += 1

                            # QBO Net Income from P&L
                            qbo_ni_row = cr
                            ws_sum.cell(cr, _DC,
                                "Net Income \u2014 QBO P&L").font = Font(name="Arial", size=10, italic=True)
                            if "P&L" in wb.sheetnames:
                                ws_pl  = wb["P&L"]
                                pl_hdr = [ws_pl.cell(_GL_HDR_R, c).value
                                          for c in range(1, ws_pl.max_column + 1)]
                                for pri in range(_GL_DATA_R, ws_pl.max_row + 1):
                                    lbl = str(ws_pl.cell(pri, _DC).value or "").strip().lower()
                                    if lbl in ("net income", "net earnings"):
                                        for ci, mk in enumerate(month_keys, _DC + 1):
                                            ml_s = month_display.get(mk, mk)
                                            for pci, ph in enumerate(pl_hdr):
                                                if str(ph or "").strip() == ml_s:
                                                    c = ws_sum.cell(cr, ci,
                                                        f"='P&L'!{get_column_letter(pci+1)}{pri}")
                                                    c.number_format = NUM_FMT
                                                    c.font = LINK_FONT
                                                    break
                                        try:
                                            ptc = pl_hdr.index("Total") + 1
                                            c = ws_sum.cell(cr, tot_col,
                                                f"='P&L'!{get_column_letter(ptc)}{pri}")
                                            c.number_format = NUM_FMT
                                            c.font = LINK_FONT
                                        except ValueError:
                                            pass
                                        break
                            cr += 1

                            # Difference row — black font, conditional fill only
                            diff_row = cr
                            ws_sum.cell(cr, _DC, "Difference (should be zero)").font = Font(name="Arial", size=10, bold=True)
                            for ci in range(_DC + 1, tot_col + 1):
                                cl = get_column_letter(ci)
                                c = ws_sum.cell(cr, ci, f"={cl}{ni_row}-{cl}{qbo_ni_row}")
                                c.number_format = NUM_FMT
                                c.font = Font(name="Arial", size=10)
                            dr = f"B{diff_row}:{get_column_letter(tot_col)}{diff_row}"
                            ws_sum.conditional_formatting.add(dr,
                                CellIsRule(operator="notEqual", formula=["0"],
                                           font=RED_FONT, fill=RED_FILL))
                            ws_sum.conditional_formatting.add(dr,
                                CellIsRule(operator="equal", formula=["0"],
                                           font=GRN_FONT, fill=GRN_FILL))
                            cr += 2

                            # ── BS Section ──────────────────────────────────
                            if "BS Balances" not in wb.sheetnames:
                                continue
                            ws_bsg  = wb["BS Balances"]
                            hdr_bsg = [ws_bsg.cell(_GL_HDR_R, c).value
                                       for c in range(1, ws_bsg.max_column + 1)]

                            bs_amt_col_i, bs_amt_col_l = _cl(hdr_bsg, "Ending Balance")
                            bs_mon_col_i, bs_mon_col_l = _cl(hdr_bsg, "Month")
                            _, bs_acct_col_l = _cl(hdr_bsg, "Account")
                            _, bs_sec_col_l = _cl(hdr_bsg, sec_label)

                            if not bs_mon_col_i or not bs_amt_col_l or not bs_sec_col_l or not bs_acct_col_l:
                                continue

                            # Collect BS month-end dates
                            bs_month_keys    = []
                            bs_month_display = {}
                            bs_month_dates   = {}
                            for ri in range(_GL_DATA_R, ws_bsg.max_row + 1):
                                mv = ws_bsg.cell(ri, bs_mon_col_i).value
                                if mv is None:
                                    continue
                                if hasattr(mv, 'strftime'):
                                    mk   = mv.strftime("%Y-%m-%d")
                                    disp = mv.strftime("%b %Y")
                                    if mk not in bs_month_dates:
                                        bs_month_dates[mk] = mv
                                else:
                                    mk   = str(mv)
                                    disp = str(mv)
                                if mk not in bs_month_keys:
                                    bs_month_keys.append(mk)
                                if mk not in bs_month_display:
                                    bs_month_display[mk] = disp
                            bs_month_keys = sorted(bs_month_keys)
                            num_bs_mo     = len(bs_month_keys)

                            # BS groups from mapping definition
                            BS_ORDER  = ["Current Assets", "Fixed Assets", "Other Assets",
                                         "Current Liabilities", "Long-term Liabilities", "Equity"]
                            bs_groups = []
                            seen_bg   = set()
                            for grp in m.get("groups", []):
                                gname = grp.get("group_name", "")
                                stmt  = grp.get("statement", "")
                                sec   = grp.get("pl_section", "") or grp.get("bs_section", "")
                                if gname and stmt == "BS" and gname not in seen_bg:
                                    seen_bg.add(gname)
                                    bs_groups.append((gname, sec))

                            bs_groups.sort(key=lambda g: BS_ORDER.index(g[1])
                                           if g[1] in BS_ORDER else 99)

                            # Title
                            ws_sum.cell(cr, _DC,
                                f"{map_name} \u2014 Balance Sheet").font = BOLD_LG
                            cr += 1

                            # BS header — col A blank, then month labels
                            ws_sum.cell(cr, _DC, "").font = HDR_FONT
                            ws_sum.cell(cr, _DC).fill = HDR_FILL
                            for ci, mk in enumerate(bs_month_keys, _DC + 1):
                                c = ws_sum.cell(cr, ci, bs_month_display.get(mk, mk))
                                c.font = HDR_FONT
                                c.fill = HDR_FILL
                                c.alignment = Alignment(horizontal="center")
                            cr += 1

                            # Build section → groups lookup
                            bs_sec_groups = {}
                            for gname, sec in bs_groups:
                                if sec not in bs_sec_groups:
                                    bs_sec_groups[sec] = []
                                bs_sec_groups[sec].append(gname)

                            ASSET_SECS = ["Current Assets", "Fixed Assets", "Other Assets"]
                            LIAB_SECS  = ["Current Liabilities", "Long-term Liabilities"]
                            EQ_SECS    = ["Equity"]

                            def _bs_date_f(mk):
                                mv = bs_month_dates.get(mk)
                                if mv and hasattr(mv, 'year'):
                                    return f"DATE({mv.year},{mv.month},{mv.day})"
                                try:
                                    d = _dt.strptime(mk[:10], "%Y-%m-%d")
                                    return f"DATE({d.year},{d.month},{d.day})"
                                except Exception:
                                    return f'"{mk}"'

                            def write_bs_block(sections_list, total_label, include_net_income=False):
                                nonlocal cr
                                sec_data_rows = []
                                for sec in sections_list:
                                    ws_sum.cell(cr, _DC, sec)
                                    sec_data_rows.append(cr)
                                    for ci, mk in enumerate(bs_month_keys, _DC + 1):
                                        date_f = _bs_date_f(mk)
                                        formula = (
                                            f"=SUMIFS('BS Balances'!${bs_amt_col_l}:${bs_amt_col_l},"
                                            f"'BS Balances'!${bs_sec_col_l}:${bs_sec_col_l},\"{sec}\","
                                            f"'BS Balances'!${bs_mon_col_l}:${bs_mon_col_l},{date_f})"
                                        )
                                        ws_sum.cell(cr, ci, formula).number_format = NUM_FMT
                                    cr += 1

                                # Net Income row (cumulative YTD from BS Balances) — Equity block only
                                ni_row = None
                                if include_net_income:
                                    ni_row = cr
                                    ws_sum.cell(cr, _DC, "Net Income").font = LINK_FONT
                                    for ci, mk in enumerate(bs_month_keys, _DC + 1):
                                        date_f = _bs_date_f(mk)
                                        formula = (
                                            f"=SUMIFS('BS Balances'!${bs_amt_col_l}:${bs_amt_col_l},"
                                            f"'BS Balances'!${bs_acct_col_l}:${bs_acct_col_l},\"Net Income\","
                                            f"'BS Balances'!${bs_mon_col_l}:${bs_mon_col_l},{date_f})"
                                        )
                                        c = ws_sum.cell(cr, ci, formula)
                                        c.number_format = NUM_FMT
                                        c.font = LINK_FONT
                                    cr += 1

                                # Total row — include Net Income for Equity
                                total_row = cr
                                ws_sum.cell(cr, _DC, total_label).font = BOLD
                                sum_rows = list(sec_data_rows)
                                if ni_row is not None:
                                    sum_rows.append(ni_row)
                                for ci in range(_DC + 1, num_bs_mo + _DC + 1):
                                    cl = get_column_letter(ci)
                                    refs = "+".join(f"{cl}{r}" for r in sum_rows)
                                    ws_sum.cell(cr, ci, f"={refs}").number_format = NUM_FMT
                                    ws_sum.cell(cr, ci).font = BOLD
                                cr += 1

                                # Balance Sheet report reference row — cross-check against QBO Balance Sheet tab
                                bs_ref_row = cr
                                ws_sum.cell(cr, _DC, f"{total_label} \u2014 QBO Balance Sheet").font = LINK_FONT
                                if "Balance Sheet" in wb.sheetnames:
                                    ws_bs_rpt = wb["Balance Sheet"]
                                    bs_rpt_hdr = [ws_bs_rpt.cell(_GL_HDR_R, c).value
                                                  for c in range(1, ws_bs_rpt.max_column + 1)]
                                    bs_match_row = None
                                    for bri in range(_GL_DATA_R, ws_bs_rpt.max_row + 1):
                                        lbl = str(ws_bs_rpt.cell(bri, _DC).value or "").strip()
                                        if lbl.lower() == total_label.lower():
                                            bs_match_row = bri
                                            break
                                    if bs_match_row:
                                        for ci, mk in enumerate(bs_month_keys, _DC + 1):
                                            ml_s = bs_month_display.get(mk, mk)
                                            for pci, ph in enumerate(bs_rpt_hdr):
                                                if str(ph or "").strip() == ml_s:
                                                    c = ws_sum.cell(cr, ci,
                                                        f"='Balance Sheet'!{get_column_letter(pci+1)}{bs_match_row}")
                                                    c.number_format = NUM_FMT
                                                    c.font = LINK_FONT
                                                    break
                                cr += 1

                                # Difference row — black font, conditional fill only
                                diff_row = cr
                                ws_sum.cell(cr, _DC, "Difference (should be zero)").font = Font(name="Arial", size=10, bold=True)
                                for ci in range(_DC + 1, num_bs_mo + _DC + 1):
                                    cl = get_column_letter(ci)
                                    c = ws_sum.cell(cr, ci, f"={cl}{total_row}-{cl}{bs_ref_row}")
                                    c.number_format = NUM_FMT
                                    c.font = Font(name="Arial", size=10)
                                dr = f"B{diff_row}:{get_column_letter(num_bs_mo + 1)}{diff_row}"
                                ws_sum.conditional_formatting.add(dr,
                                    CellIsRule(operator="notEqual", formula=["0"], font=RED_FONT, fill=RED_FILL))
                                ws_sum.conditional_formatting.add(dr,
                                    CellIsRule(operator="equal", formula=["0"], font=GRN_FONT, fill=GRN_FILL))
                                cr += 2
                                return total_row

                            ta_row  = write_bs_block(ASSET_SECS, "Total Assets")
                            tl_row  = write_bs_block(LIAB_SECS, "Total Liabilities")
                            te_row  = write_bs_block(EQ_SECS, "Total Equity", include_net_income=True)

                            # Total Liabilities & Equity
                            tle_row = cr
                            ws_sum.cell(cr, _DC, "Total Liabilities & Equity").font = BOLD
                            for ci in range(_DC + 1, num_bs_mo + _DC + 1):
                                cl = get_column_letter(ci)
                                ws_sum.cell(cr, ci, f"={cl}{tl_row}+{cl}{te_row}").number_format = NUM_FMT
                                ws_sum.cell(cr, ci).font = BOLD
                            cr += 2

                            # Balance check — Total Assets vs Total Liabilities & Equity
                            ws_sum.cell(cr, _DC, "Total Assets").font = BOLD
                            for ci in range(_DC + 1, num_bs_mo + _DC + 1):
                                cl = get_column_letter(ci)
                                ws_sum.cell(cr, ci, f"={cl}{ta_row}").number_format = NUM_FMT
                                ws_sum.cell(cr, ci).font = BOLD
                            cr += 1

                            ws_sum.cell(cr, _DC, "Total Liabilities & Equity").font = BOLD
                            for ci in range(_DC + 1, num_bs_mo + _DC + 1):
                                cl = get_column_letter(ci)
                                ws_sum.cell(cr, ci, f"={cl}{tle_row}").number_format = NUM_FMT
                                ws_sum.cell(cr, ci).font = BOLD
                            cr += 1

                            bal_diff_row = cr
                            ws_sum.cell(cr, _DC, "Difference (should be zero)").font = Font(name="Arial", size=10, bold=True)
                            for ci in range(_DC + 1, num_bs_mo + _DC + 1):
                                cl = get_column_letter(ci)
                                c = ws_sum.cell(cr, ci, f"={cl}{ta_row}-{cl}{tle_row}")
                                c.number_format = NUM_FMT
                                c.font = Font(name="Arial", size=10)
                            bdr = f"B{bal_diff_row}:{get_column_letter(num_bs_mo + 1)}{bal_diff_row}"
                            ws_sum.conditional_formatting.add(bdr,
                                CellIsRule(operator="notEqual", formula=["0"], font=RED_FONT, fill=RED_FILL))
                            ws_sum.conditional_formatting.add(bdr,
                                CellIsRule(operator="equal", formula=["0"], font=GRN_FONT, fill=GRN_FILL))
                            cr += 2

                            # ── Mapped P&L and BS tabs ──────────────────────
                            try:
                                from openpyxl.styles import Border, Side
                                _THIN  = Side(style="thin")
                                _DBLE  = Side(style="double")
                                _MAPPL = Font(name="Arial", size=10)
                                _MAPBD = Font(name="Arial", size=10, bold=True)
                                _GRAY  = PatternFill("solid", fgColor="F2F2F2")
                                _NOFLL = PatternFill(fill_type=None)
                                _ITALF = Font(name="Arial", size=10, italic=True)

                                def _is_date_formula(mk):
                                    mv = month_dates.get(mk)
                                    if mv and hasattr(mv, 'year'):
                                        return f"DATE({mv.year},{mv.month},{mv.day})"
                                    try:
                                        d = _dt.strptime(mk, "%Y-%m")
                                        last = _cal.monthrange(d.year, d.month)[1]
                                        return f"DATE({d.year},{d.month},{last})"
                                    except Exception:
                                        return f'"{mk}"'

                                def _diff_cfmt(ws, r, last_ci):
                                    dr = f"B{r}:{get_column_letter(last_ci)}{r}"
                                    ws.conditional_formatting.add(dr,
                                        CellIsRule(operator="notEqual", formula=["0"],
                                                   font=RED_FONT, fill=RED_FILL))
                                    ws.conditional_formatting.add(dr,
                                        CellIsRule(operator="equal", formula=["0"],
                                                   font=GRN_FONT, fill=GRN_FILL))

                                # ── {Map Name} P&L ─────────────────────────
                                _is_sg = {}
                                for gn, sc in is_groups:
                                    _is_sg.setdefault(sc, []).append(gn)

                                if _is_sg and is_grp_col_l:
                                    _pl_tab = f"{map_name} P&L"
                                    if _pl_tab in wb.sheetnames:
                                        del wb[_pl_tab]
                                    wpl = wb.create_sheet(_pl_tab)
                                    wpl.sheet_view.showGridLines = False
                                    wpl.cell(1, 1, company_name).font = Font(name="Arial", size=10, bold=True)
                                    wpl.cell(2, 1, f"{map_name} \u2014 Income Statement").font = Font(name="Arial", size=10)

                                    wpl.cell(5, _DC, "Account").font = HDR_FONT
                                    wpl.cell(5, _DC).fill = HDR_FILL
                                    wpl.cell(5, _DC).alignment = Alignment(horizontal="left")
                                    for ci, mk in enumerate(month_keys, _DC + 1):
                                        c = wpl.cell(5, ci, month_display.get(mk, mk))
                                        c.font = HDR_FONT; c.fill = HDR_FILL
                                        c.alignment = Alignment(horizontal="center")
                                    wpl.cell(5, tot_col, "Total").font = HDR_FONT
                                    wpl.cell(5, tot_col).fill = HDR_FILL
                                    wpl.cell(5, tot_col).alignment = Alignment(horizontal="center")

                                    pr = 6
                                    rr = {}

                                    def _write_sec(sec):
                                        nonlocal pr
                                        grps = _is_sg.get(sec)
                                        if not grps:
                                            return
                                        c = wpl.cell(pr, _DC, sec)
                                        c.font = _MAPBD; c.fill = _GRAY
                                        for ci in range(_DC + 1, tot_col + 1):
                                            wpl.cell(pr, ci).font = _MAPPL
                                            wpl.cell(pr, ci).fill = _GRAY
                                        pr += 1
                                        grp_rows = []
                                        for gn in grps:
                                            wpl.cell(pr, _DC, gn).font = _MAPPL
                                            wpl.cell(pr, _DC).alignment = Alignment(indent=1)
                                            grp_rows.append(pr)
                                            for ci, mk in enumerate(month_keys, _DC + 1):
                                                df = _is_date_formula(mk)
                                                c = wpl.cell(pr, ci,
                                                    f"=SUMIFS('IS GL Summary'!${is_amt_col_l}:${is_amt_col_l},"
                                                    f"'IS GL Summary'!${is_grp_col_l}:${is_grp_col_l},\"{gn}\","
                                                    f"'IS GL Summary'!${is_mon_col_l}:${is_mon_col_l},{df})")
                                                c.number_format = NUM_FMT
                                            sl = get_column_letter(_DC + 1)
                                            el = get_column_letter(1 + num_mo)
                                            wpl.cell(pr, tot_col, f"=SUM({sl}{pr}:{el}{pr})").number_format = NUM_FMT
                                            pr += 1
                                        wpl.cell(pr, _DC, f"Total {sec}").font = _MAPBD
                                        for ci in range(_DC + 1, tot_col + 1):
                                            cl = get_column_letter(ci)
                                            refs = "+".join(f"{cl}{r}" for r in grp_rows)
                                            c = wpl.cell(pr, ci, f"={refs}")
                                            c.number_format = NUM_FMT; c.font = _MAPBD
                                            c.border = Border(top=_THIN)
                                        rr[sec] = pr
                                        pr += 2

                                    def _calc(label, plus, minus, top_b=False, dbl_b=False):
                                        nonlocal pr
                                        lc = wpl.cell(pr, _DC, label)
                                        lc.font = _MAPBD
                                        if top_b or dbl_b:
                                            lc.border = Border(
                                                top=_THIN if top_b else None,
                                                bottom=_DBLE if dbl_b else None)
                                        for ci in range(_DC + 1, tot_col + 1):
                                            cl = get_column_letter(ci)
                                            pp = [f"+{cl}{rr[k]}" for k in plus if k in rr]
                                            mm = [f"-{cl}{rr[k]}" for k in minus if k in rr]
                                            f = "=" + "".join(pp + mm) if (pp or mm) else "=0"
                                            c = wpl.cell(pr, ci, f)
                                            c.number_format = NUM_FMT; c.font = _MAPBD
                                            if top_b or dbl_b:
                                                c.border = Border(
                                                    top=_THIN if top_b else None,
                                                    bottom=_DBLE if dbl_b else None)
                                        rv = pr; pr += 1; return rv

                                    REV = ["Revenue"]
                                    COS = ["COS", "Cost of Goods Sold"]
                                    SM  = ["Sales & Marketing"]
                                    OPX = ["Operating Expenses"]
                                    OI  = ["Other Income"]
                                    OE  = ["Other Expense", "Other"]

                                    # Revenue + COS sections
                                    _write_sec("Revenue")
                                    _write_sec("COS")
                                    _write_sec("Cost of Goods Sold")

                                    # Gross Profit immediately after COS
                                    gp = _calc("Gross Profit", REV, COS)
                                    rr["_GP"] = gp
                                    pr += 1  # blank row before next section

                                    # Remaining sections
                                    _write_sec("Sales & Marketing")
                                    _write_sec("Operating Expenses")
                                    _write_sec("Other Income")
                                    _write_sec("Other Expense")
                                    _write_sec("Other")

                                    if any(k in rr for k in SM):
                                        cm = _calc("Contribution Margin", ["_GP"], SM)
                                        rr["_CM"] = cm
                                    else:
                                        rr["_CM"] = gp
                                    noi = _calc("Net Operating Income", ["_CM"], OPX)
                                    rr["_NOI"] = noi
                                    ni_pl_row = _calc("Net Income", ["_NOI"] + OI, OE,
                                                      top_b=True, dbl_b=True)

                                    # ── P&L validation: Net Income cross-check ──
                                    pr += 1
                                    qbo_ni_r = pr
                                    wpl.cell(pr, _DC, "Net Income \u2014 QBO P&L").font = Font(
                                        name="Arial", size=10, italic=True, color="276221")
                                    if "P&L" in wb.sheetnames:
                                        ws_pl_src = wb["P&L"]
                                        pl_src_hdr = [ws_pl_src.cell(_GL_HDR_R, c).value
                                                      for c in range(1, ws_pl_src.max_column + 1)]
                                        for pri in range(_GL_DATA_R, ws_pl_src.max_row + 1):
                                            lbl = str(ws_pl_src.cell(pri, _DC).value or "").strip().lower()
                                            if lbl in ("net income", "net earnings"):
                                                for ci, mk in enumerate(month_keys, _DC + 1):
                                                    ml_s = month_display.get(mk, mk)
                                                    for pci, ph in enumerate(pl_src_hdr):
                                                        if str(ph or "").strip() == ml_s:
                                                            c = wpl.cell(pr, ci,
                                                                f"='P&L'!{get_column_letter(pci+1)}{pri}")
                                                            c.number_format = NUM_FMT
                                                            c.font = LINK_FONT
                                                            break
                                                try:
                                                    ptc = pl_src_hdr.index("Total") + 1
                                                    c = wpl.cell(pr, tot_col,
                                                        f"='P&L'!{get_column_letter(ptc)}{pri}")
                                                    c.number_format = NUM_FMT
                                                    c.font = LINK_FONT
                                                except ValueError:
                                                    pass
                                                break
                                    pr += 1

                                    diff_r = pr
                                    wpl.cell(pr, _DC, "Difference (should be zero)").font = _MAPBD
                                    for ci in range(_DC + 1, tot_col + 1):
                                        cl = get_column_letter(ci)
                                        c = wpl.cell(pr, ci, f"={cl}{ni_pl_row}-{cl}{qbo_ni_r}")
                                        c.number_format = NUM_FMT
                                        c.font = Font(name="Arial", size=10)
                                    _diff_cfmt(wpl, diff_r, tot_col)
                                    pr += 1

                                    wpl.column_dimensions["A"].width = 0.63
                                    _max_b = 0
                                    for _r in range(5, wpl.max_row + 1):
                                        _v = wpl.cell(_r, _DC).value
                                        if _v is not None:
                                            _max_b = max(_max_b, len(str(_v)))
                                    wpl.column_dimensions[get_column_letter(_DC)].width = min(max(_max_b + 4, 30), 50)
                                    for ci in range(_DC + 1, tot_col + 1):
                                        wpl.column_dimensions[get_column_letter(ci)].width = 14
                                    wpl.freeze_panes = "A6"

                                # ── {Map Name} BS ──────────────────────────
                                _, _bs_grp_l = _cl(hdr_bsg, grp_label)

                                if bs_groups and _bs_grp_l:
                                    _bs_tab = f"{map_name} BS"
                                    if _bs_tab in wb.sheetnames:
                                        del wb[_bs_tab]
                                    wbs = wb.create_sheet(_bs_tab)
                                    wbs.sheet_view.showGridLines = False
                                    wbs.cell(1, 1, company_name).font = Font(name="Arial", size=10, bold=True)
                                    wbs.cell(2, 1, f"{map_name} \u2014 Balance Sheet").font = Font(name="Arial", size=10)

                                    wbs.cell(5, _DC, "Account").font = HDR_FONT
                                    wbs.cell(5, _DC).fill = HDR_FILL
                                    wbs.cell(5, _DC).alignment = Alignment(horizontal="left")
                                    for ci, mk in enumerate(bs_month_keys, _DC + 1):
                                        c = wbs.cell(5, ci, bs_month_display.get(mk, mk))
                                        c.font = HDR_FONT; c.fill = HDR_FILL
                                        c.alignment = Alignment(horizontal="center")

                                    br = 6
                                    brr = {}

                                    def _bs_sec_block(sec, sec_label_display=None):
                                        nonlocal br
                                        grps = bs_sec_groups.get(sec, [])
                                        if not grps:
                                            sec_low = sec.lower()
                                            for k, v in bs_sec_groups.items():
                                                if k.lower() == sec_low:
                                                    grps = v
                                                    break
                                        lbl  = sec_label_display or sec

                                        # Section header — gray fill, bold black
                                        c = wbs.cell(br, _DC, lbl)
                                        c.font = _MAPBD; c.fill = _GRAY
                                        for ci in range(_DC + 1, num_bs_mo + _DC + 1):
                                            wbs.cell(br, ci).font = _MAPPL
                                            wbs.cell(br, ci).fill = _GRAY
                                        br += 1

                                        grp_rows = []
                                        for gn in grps:
                                            wbs.cell(br, _DC, gn).font = _MAPPL
                                            wbs.cell(br, _DC).alignment = Alignment(indent=1)
                                            grp_rows.append(br)
                                            for ci, mk in enumerate(bs_month_keys, _DC + 1):
                                                df = _bs_date_f(mk)
                                                wbs.cell(br, ci,
                                                    f"=SUMIFS('BS Balances'!${bs_amt_col_l}:${bs_amt_col_l},"
                                                    f"'BS Balances'!${_bs_grp_l}:${_bs_grp_l},\"{gn}\","
                                                    f"'BS Balances'!${bs_mon_col_l}:${bs_mon_col_l},{df})"
                                                ).number_format = NUM_FMT
                                            br += 1

                                        if not grp_rows:
                                            wbs.cell(br, _DC, "(none)").font = _MAPPL
                                            wbs.cell(br, _DC).alignment = Alignment(indent=1)
                                            grp_rows.append(br)
                                            for ci in range(_DC + 1, num_bs_mo + _DC + 1):
                                                wbs.cell(br, ci, 0).number_format = NUM_FMT
                                            br += 1

                                        return grp_rows

                                    def _bs_total(label, row_lists, top_b=False, dbl_b=False):
                                        nonlocal br
                                        all_rows = [r for rl in row_lists for r in rl]
                                        lc = wbs.cell(br, _DC, label)
                                        lc.font = _MAPBD
                                        if top_b or dbl_b:
                                            lc.border = Border(
                                                top=_THIN if top_b else None,
                                                bottom=_DBLE if dbl_b else None)
                                        for ci in range(_DC + 1, num_bs_mo + _DC + 1):
                                            cl = get_column_letter(ci)
                                            refs = "+".join(f"{cl}{r}" for r in all_rows) if all_rows else "0"
                                            c = wbs.cell(br, ci, f"={refs}")
                                            c.number_format = NUM_FMT; c.font = _MAPBD
                                            if top_b or dbl_b:
                                                c.border = Border(
                                                    top=_THIN if top_b else None,
                                                    bottom=_DBLE if dbl_b else None)
                                        rv = br; br += 1; return rv

                                    def _bs_ref_total(label, ref_rows, top_b=False, dbl_b=False):
                                        nonlocal br
                                        lc = wbs.cell(br, _DC, label)
                                        lc.font = _MAPBD
                                        if top_b or dbl_b:
                                            lc.border = Border(
                                                top=_THIN if top_b else None,
                                                bottom=_DBLE if dbl_b else None)
                                        for ci in range(_DC + 1, num_bs_mo + _DC + 1):
                                            cl = get_column_letter(ci)
                                            refs = "+".join(f"{cl}{r}" for r in ref_rows) if ref_rows else "0"
                                            c = wbs.cell(br, ci, f"={refs}")
                                            c.number_format = NUM_FMT; c.font = _MAPBD
                                            if top_b or dbl_b:
                                                c.border = Border(
                                                    top=_THIN if top_b else None,
                                                    bottom=_DBLE if dbl_b else None)
                                        rv = br; br += 1; return rv

                                    # Assets
                                    ca_rows = _bs_sec_block("Current Assets")
                                    fa_rows = _bs_sec_block("Fixed Assets")
                                    oa_rows = _bs_sec_block("Other Assets")
                                    ta_r = _bs_total("Total Assets", [ca_rows, fa_rows, oa_rows],
                                                     top_b=True, dbl_b=True)
                                    brr["TA"] = ta_r
                                    br += 1

                                    # Liabilities
                                    cl_rows = _bs_sec_block("Current Liabilities")
                                    ll_rows = _bs_sec_block("Long-term Liabilities")
                                    tl_r = _bs_total("Total Liabilities", [cl_rows, ll_rows], top_b=True)
                                    brr["TL"] = tl_r
                                    br += 1

                                    # Equity
                                    eq_rows = _bs_sec_block("Equity")

                                    ni_r = br
                                    wbs.cell(br, _DC, "Net Income").font = Font(name="Arial", size=10, color="276221")
                                    wbs.cell(br, _DC).alignment = Alignment(indent=1)
                                    for ci, mk in enumerate(bs_month_keys, _DC + 1):
                                        df = _bs_date_f(mk)
                                        c = wbs.cell(br, ci,
                                            f"=SUMIFS('BS Balances'!${bs_amt_col_l}:${bs_amt_col_l},"
                                            f"'BS Balances'!${bs_acct_col_l}:${bs_acct_col_l},\"Net Income\","
                                            f"'BS Balances'!${bs_mon_col_l}:${bs_mon_col_l},{df})")
                                        c.number_format = NUM_FMT
                                        c.font = Font(name="Arial", size=10, color="276221")
                                    br += 1

                                    te_r = _bs_total("Total Equity", [eq_rows, [ni_r]], top_b=True)
                                    brr["TE"] = te_r
                                    br += 1

                                    tle_r = br
                                    _bs_ref_total("Total Liabilities & Equity",
                                                  [brr["TL"], brr["TE"]], top_b=True, dbl_b=True)

                                    # ── BS validation ──────────────────────────
                                    # Balance check: TA vs TL&E
                                    br += 1
                                    _last_bs_ci = num_bs_mo + _DC

                                    wbs.cell(br, _DC, "Total Assets").font = _MAPBD
                                    for ci in range(_DC + 1, num_bs_mo + _DC + 1):
                                        cl = get_column_letter(ci)
                                        c = wbs.cell(br, ci, f"={cl}{brr['TA']}")
                                        c.number_format = NUM_FMT; c.font = _MAPBD
                                    br += 1

                                    wbs.cell(br, _DC, "Total Liabilities & Equity").font = _MAPBD
                                    for ci in range(_DC + 1, num_bs_mo + _DC + 1):
                                        cl = get_column_letter(ci)
                                        c = wbs.cell(br, ci, f"={cl}{tle_r}")
                                        c.number_format = NUM_FMT; c.font = _MAPBD
                                    br += 1

                                    d1_r = br
                                    wbs.cell(br, _DC, "Difference (should be zero)").font = _MAPBD
                                    for ci in range(_DC + 1, num_bs_mo + _DC + 1):
                                        cl = get_column_letter(ci)
                                        c = wbs.cell(br, ci, f"={cl}{brr['TA']}-{cl}{tle_r}")
                                        c.number_format = NUM_FMT
                                        c.font = Font(name="Arial", size=10)
                                    _diff_cfmt(wbs, d1_r, _last_bs_ci)
                                    br += 2

                                    # QBO Balance Sheet cross-check
                                    _bs_check_labels = [
                                        ("Total Assets", brr["TA"]),
                                        ("Total Liabilities", brr["TL"]),
                                        ("Total Equity", brr["TE"]),
                                    ]
                                    if "Balance Sheet" in wb.sheetnames:
                                        ws_bs_src = wb["Balance Sheet"]
                                        bs_src_hdr = [ws_bs_src.cell(_GL_HDR_R, c).value
                                                      for c in range(1, ws_bs_src.max_column + 1)]

                                        for chk_label, mapped_row in _bs_check_labels:
                                            bs_match_r = None
                                            for bri in range(_GL_DATA_R, ws_bs_src.max_row + 1):
                                                lbl = str(ws_bs_src.cell(bri, _DC).value or "").strip()
                                                if lbl.lower() == chk_label.lower():
                                                    bs_match_r = bri
                                                    break

                                            qbo_r = br
                                            wbs.cell(br, _DC,
                                                f"{chk_label} \u2014 QBO Balance Sheet").font = Font(
                                                    name="Arial", size=10, italic=True, color="276221")
                                            if bs_match_r:
                                                for ci, mk in enumerate(bs_month_keys, _DC + 1):
                                                    ml_s = bs_month_display.get(mk, mk)
                                                    for pci, ph in enumerate(bs_src_hdr):
                                                        if str(ph or "").strip() == ml_s:
                                                            c = wbs.cell(br, ci,
                                                                f"='Balance Sheet'!{get_column_letter(pci+1)}{bs_match_r}")
                                                            c.number_format = NUM_FMT
                                                            c.font = LINK_FONT
                                                            break
                                            br += 1

                                            d_r = br
                                            wbs.cell(br, _DC, "Difference (should be zero)").font = _MAPBD
                                            for ci in range(_DC + 1, num_bs_mo + _DC + 1):
                                                cl = get_column_letter(ci)
                                                c = wbs.cell(br, ci,
                                                    f"={cl}{mapped_row}-{cl}{qbo_r}")
                                                c.number_format = NUM_FMT
                                                c.font = Font(name="Arial", size=10)
                                            _diff_cfmt(wbs, d_r, _last_bs_ci)
                                            br += 2

                                    wbs.column_dimensions["A"].width = 0.63
                                    _max_b = 0
                                    for _r in range(5, wbs.max_row + 1):
                                        _v = wbs.cell(_r, _DC).value
                                        if _v is not None:
                                            _max_b = max(_max_b, len(str(_v)))
                                    wbs.column_dimensions[get_column_letter(_DC)].width = min(max(_max_b + 4, 30), 50)
                                    for ci in range(_DC + 1, num_bs_mo + _DC + 1):
                                        wbs.column_dimensions[get_column_letter(ci)].width = 14
                                    wbs.freeze_panes = "A6"

                            except Exception as _mpt_e:
                                import traceback
                                logger.warning(f"Mapped P&L/BS tabs failed for '{map_name}': {_mpt_e}\n{traceback.format_exc()}")

                        # Column widths for Map Summary
                        ws_sum.column_dimensions["A"].width = 0.63
                        _max_b = 0
                        for _r in range(5, ws_sum.max_row + 1):
                            _v = ws_sum.cell(_r, _DC).value
                            if _v is not None:
                                _max_b = max(_max_b, len(str(_v)))
                        ws_sum.column_dimensions[get_column_letter(_DC)].width = min(max(_max_b + 4, 30), 50)
                        for ci in range(_DC + 1, ws_sum.max_column + 1):
                            ws_sum.column_dimensions[get_column_letter(ci)].width = 14
                        ws_sum.freeze_panes = "A6"

                        for _ws in wb.worksheets:
                            for _row in _ws.iter_rows(min_row=1, max_row=_ws.max_row, max_col=_ws.max_column):
                                for _c in _row:
                                    if _c.font and (_c.font.name != "Arial" or _c.font.size != 10):
                                        _c.font = Font(
                                            name="Arial", size=10,
                                            bold=_c.font.bold, italic=_c.font.italic,
                                            color=_c.font.color, underline=_c.font.underline,
                                            strikethrough=_c.font.strikethrough,
                                        )

                        wb.save(file_path)
                        progress_fn("  Mapping columns and Map Summary written.")

                except Exception as e:
                    import traceback
                    logger.error(f"Mapping error: {e}\n{traceback.format_exc()}")
                    progress_fn(f"  WARNING: Mapping failed — {e}")

            # Add portal flat tabs if requested
            if include_portal_data:
                try:
                    progress_fn("  Building portal data tabs...")
                    import openpyxl as _ox_p
                    from backend.portal_prep import build_portal_flat_tabs
                    wb_p = _ox_p.load_workbook(file_path)

                    # Read IS GL Summary and BS GL Summary rows from the workbook
                    def _read_tab_rows(wb, tab_name):
                        if tab_name not in wb.sheetnames:
                            return []
                        ws = wb[tab_name]
                        return [[ws.cell(r, c).value for c in range(1, ws.max_column+1)]
                                for r in range(5, ws.max_row+1)]

                    is_sum = _read_tab_rows(wb_p, "IS GL Summary")
                    bs_bal = _read_tab_rows(wb_p, "BS Balances")
                    p_is, p_bs = build_portal_flat_tabs(is_sum, bs_bal)

                    if p_is:
                        from openpyxl.styles import Font as _Fp, PatternFill as _PFp, Alignment as _Alp
                        from openpyxl.utils import get_column_letter as _gclp
                        from datetime import datetime as _dtp, date as _datep
                        _PORTAL_NAMES = {"Portal_IS_Flat": "Portal Income Statement",
                                         "Portal_BS_Flat": "Portal Balance Sheet"}
                        for tab_name, rows in [("Portal_IS_Flat", p_is), ("Portal_BS_Flat", p_bs)]:
                            if not rows: continue
                            ws = wb_p.create_sheet(tab_name)
                            ws.sheet_view.showGridLines = False
                            ws.cell(1, 1, company_name).font = _Fp(name="Arial", size=10, bold=True)
                            ws.cell(2, 1, _PORTAL_NAMES.get(tab_name, tab_name)).font = _Fp(name="Arial", size=10)
                            hf = _Fp(name="Arial", size=10, bold=True, color="FFFFFF")
                            hb = _PFp("solid", fgColor="337E8D")
                            pf = _Fp(name="Arial", size=10)
                            ws.column_dimensions["A"].width = 0.63
                            for ci, v in enumerate(rows[0], 2):
                                c = ws.cell(5, ci, v); c.font = hf; c.fill = hb
                            for ri, row in enumerate(rows[1:], 6):
                                for ci, v in enumerate(row, 2):
                                    if isinstance(v, _dtp): v = v.date()
                                    c = ws.cell(ri, ci, v); c.font = pf
                                    if isinstance(v, _datep):
                                        c.number_format = "M/D/YYYY"
                                    elif isinstance(v, (int, float)):
                                        c.number_format = '#,##0.00_);(#,##0.00);"-"??;@'
                            # Autofit column widths (skip buffer col A)
                            for col_cells in ws.columns:
                                col_idx = col_cells[0].column
                                if col_idx < 2:
                                    continue
                                mx = 0
                                cl = _gclp(col_idx)
                                for cell in col_cells:
                                    if cell.row < 5:
                                        continue
                                    try:
                                        cl2 = len(str(cell.value)) if cell.value is not None else 0
                                        if cl2 > mx: mx = cl2
                                    except: pass
                                ws.column_dimensions[cl].width = min(mx + 4, 60)
                            ws.freeze_panes = "A6"

                    for _ws in wb_p.worksheets:
                        for _row in _ws.iter_rows(min_row=1, max_row=_ws.max_row, max_col=_ws.max_column):
                            for _c in _row:
                                if _c.font and (_c.font.name != "Arial" or _c.font.size != 10):
                                    _c.font = Font(
                                        name="Arial", size=10,
                                        bold=_c.font.bold, italic=_c.font.italic,
                                        color=_c.font.color, underline=_c.font.underline,
                                        strikethrough=_c.font.strikethrough,
                                    )

                    wb_p.save(file_path)
                    progress_fn("  Portal data tabs added.")
                except Exception as pe:
                    logger.warning(f"Portal tab build failed: {pe}")
                    progress_fn(f"  WARNING: Portal tabs failed — {pe}")

            # ── Final restructuring: Summary tab, dividers, tab order ──
            try:
                progress_fn("  Finalizing workbook structure...")
                import openpyxl as _ox_fin
                from openpyxl.styles import Font as _Ff, PatternFill as _PFf, Alignment as _Alf
                from openpyxl.utils import get_column_letter as _gclf
                from pathlib import Path as _Pf

                wb_fin = _ox_fin.load_workbook(file_path)

                # ── Rename tabs ──
                if "Validation" in wb_fin.sheetnames:
                    wb_fin["Validation"].title = "GL Summary Validation"

                # Derive map names from Mapped P&L tabs
                _map_names = []
                for sn in wb_fin.sheetnames:
                    if sn.endswith(" P&L") and sn != "P&L":
                        _map_names.append(sn[:-4])

                if "Map Summary" in wb_fin.sheetnames and _map_names:
                    wb_fin["Map Summary"].title = f"{_map_names[0]} Validation"

                # ── Divider helper ──
                _DIV_COLOR = "07393C"

                def _make_divider(name):
                    ws = wb_fin.create_sheet(name)
                    ws.sheet_view.showGridLines = False
                    ws.sheet_properties.tabColor = _DIV_COLOR

                # ── Build ordered tab list ──
                ordered = []

                # Summary (will be created below)
                ordered.append("Summary")

                # QBO Reports section
                qbo_tabs = []
                for t in ["P&L", "Balance Sheet", "AR Aging", "AP Aging"]:
                    if t in wb_fin.sheetnames:
                        qbo_tabs.append(t)
                if qbo_tabs:
                    ordered.append("QBO Reports")
                    _make_divider("QBO Reports")
                    ordered.extend(qbo_tabs)

                # Mapped Reports section
                mapped_tabs = []
                for mn in _map_names:
                    for suffix in [" P&L", " BS"]:
                        tn = mn + suffix
                        if tn in wb_fin.sheetnames:
                            mapped_tabs.append(tn)
                if mapped_tabs:
                    ordered.append("Mapped Reports")
                    _make_divider("Mapped Reports")
                    ordered.extend(mapped_tabs)

                # Flat Files section
                flat_tabs = []
                for t in ["IS GL Summary", "BS GL Summary", "BS Balances",
                           "IS GL Detail", "BS GL Detail"]:
                    if t in wb_fin.sheetnames:
                        flat_tabs.append(t)
                if flat_tabs:
                    ordered.append("Flat Files")
                    _make_divider("Flat Files")
                    ordered.extend(flat_tabs)

                # Data Validation section
                val_tabs = []
                if "GL Summary Validation" in wb_fin.sheetnames:
                    val_tabs.append("GL Summary Validation")
                for mn in _map_names:
                    vn = f"{mn} Validation"
                    if vn in wb_fin.sheetnames:
                        val_tabs.append(vn)
                if val_tabs:
                    ordered.append("Data Validation")
                    _make_divider("Data Validation")
                    ordered.extend(val_tabs)

                # Mapping Reference (after validation, before portal)
                if "Mapping Reference" in wb_fin.sheetnames:
                    ordered.append("Mapping Reference")

                # Portal Data section
                portal_tabs = []
                for t in ["Portal_IS_Flat", "Portal_BS_Flat"]:
                    if t in wb_fin.sheetnames:
                        portal_tabs.append(t)
                if portal_tabs:
                    ordered.append("Portal Data")
                    _make_divider("Portal Data")
                    ordered.extend(portal_tabs)

                # Add any remaining tabs not in the ordered list
                for sn in wb_fin.sheetnames:
                    if sn not in ordered and sn != "Summary":
                        ordered.append(sn)

                # ── Create Summary tab ──
                ws_s = wb_fin.create_sheet("Summary", 0)
                ws_s.sheet_view.showGridLines = False
                ws_s.sheet_properties.tabColor = "C97D60"
                ws_s.column_dimensions["A"].width = 0.63

                _SF  = _Ff(name="Arial", size=10)
                _SFB = _Ff(name="Arial", size=10, bold=True)
                _SFG = _Ff(name="Arial", size=10, color="5A6B6D")
                _SFI = _Ff(name="Arial", size=10, italic=True, color="5A6B6D")
                _SHF = _Ff(name="Arial", size=10, bold=True, color="FFFFFF")
                _SHB = _PFf("solid", fgColor="337E8D")
                _SLK = _Ff(name="Arial", size=10, color="0563C1", underline="single")
                _NUM = '#,##0.00_);(#,##0.00);"-"??;@'

                def _sec_bar(r, text):
                    for ci in range(2, 6):
                        ws_s.cell(r, ci).fill = _SHB
                    ws_s.cell(r, 2, text).font = _SHF

                # Row 1: company name (A1 overflows) + logo at D1
                ws_s.cell(1, 1, company_name).font = _Ff(name="Arial", size=14, bold=True, color="07393C")
                try:
                    import os as _os_logo
                    from openpyxl.drawing.image import Image as _XlImg
                    _logo_candidates = [
                        _os_logo.path.join(_os_logo.path.dirname(_os_logo.path.abspath(__file__)), "assets", "Logo-F23-transparent.png"),
                        _os_logo.path.join(_os_logo.path.dirname(__file__), "assets", "Logo-F23-transparent.png"),
                        _os_logo.path.join("backend", "assets", "Logo-F23-transparent.png"),
                    ]
                    _logo_path = None
                    for _lp in _logo_candidates:
                        if _os_logo.path.exists(_lp):
                            _logo_path = _lp
                            break
                    if _logo_path:
                        img = _XlImg(_logo_path)
                        _ratio = img.height / max(img.width, 1)
                        img.width = 200; img.height = int(200 * _ratio)
                        ws_s.add_image(img, "D1")
                except Exception as _logo_err:
                    logger.warning(f"Could not load logo: {_logo_err}")

                ws_s.cell(2, 1, "Summary").font = _SF
                ws_s.cell(3, 1, "Acorn by Oakbridge Finance").font = _SFI

                # Row 5: Report Summary bar
                _sec_bar(5, "Report Summary")
                from datetime import datetime as _dtf
                ws_s.cell(6, 2, f"Report Period: {start_date} \u2014 {end_date}").font = _SFG
                try:
                    from zoneinfo import ZoneInfo
                    _now = _dtf.now(ZoneInfo("America/Chicago"))
                except Exception:
                    _now = _dtf.now()
                try:
                    _time_s = _now.strftime("%-I:%M %p")
                except ValueError:
                    _time_s = _now.strftime("%I:%M %p").lstrip("0")
                ws_s.cell(7, 2, f"Generated: {_now.strftime('%B %d, %Y')} at {_time_s}").font = _SFG

                # Row 9: Validation Summary bar
                _sec_bar(9, "Validation Summary")

                ws_s.cell(10, 2, "QBO Reports").font = _SFB

                _val_ref_col = "D"
                _SGR = _Ff(name="Arial", size=10, color="276221")
                if "GL Summary Validation" in wb_fin.sheetnames:
                    ws_s.cell(11, 2, "Overall Result").font = _SF
                    ws_s.cell(11, 4, f"='GL Summary Validation'!{_val_ref_col}6").font = _SGR
                    ws_s.cell(12, 2, "Total Accounts Checked").font = _SF
                    ws_s.cell(12, 4, f"='GL Summary Validation'!{_val_ref_col}7").font = _SF
                    ws_s.cell(13, 2, "Matched").font = _SF
                    ws_s.cell(13, 4, f"='GL Summary Validation'!{_val_ref_col}8").font = _SF
                    ws_s.cell(14, 2, "Differences Found").font = _SF
                    ws_s.cell(14, 4, f"='GL Summary Validation'!{_val_ref_col}9").font = _SF
                    ws_s.cell(15, 2, "Missing from GL").font = _SF
                    ws_s.cell(15, 4, f"='GL Summary Validation'!{_val_ref_col}10").font = _SF

                _sr = 17
                for mn in _map_names:
                    pl_tn = f"{mn} P&L"
                    bs_tn = f"{mn} BS"
                    ws_s.cell(_sr, 2, mn).font = _SFB
                    _sr += 1

                    _pl_diff_r = None
                    _bs_diff_rs = {}
                    if pl_tn in wb_fin.sheetnames:
                        ws_mp = wb_fin[pl_tn]
                        for r in range(6, ws_mp.max_row + 1):
                            v = str(ws_mp.cell(r, 2).value or "")
                            if v.startswith("Difference"):
                                _pl_diff_r = r; break

                    if bs_tn in wb_fin.sheetnames:
                        ws_mb = wb_fin[bs_tn]
                        _prev_qbo = None
                        for r in range(6, ws_mb.max_row + 1):
                            v = str(ws_mb.cell(r, 2).value or "")
                            if "QBO" in v:
                                for bl in ["Total Assets", "Total Liabilities", "Total Equity"]:
                                    if bl in v:
                                        _prev_qbo = bl
                            elif v.startswith("Difference") and _prev_qbo:
                                _bs_diff_rs[_prev_qbo] = r
                                _prev_qbo = None
                            elif v.startswith("Difference") and not _prev_qbo and "Balance Check" not in _bs_diff_rs:
                                _bs_diff_rs["Balance Check"] = r

                    _checks = [
                        ("Net Income", _pl_diff_r, pl_tn),
                        ("Total Assets", _bs_diff_rs.get("Total Assets"), bs_tn),
                        ("Total Liabilities", _bs_diff_rs.get("Total Liabilities"), bs_tn),
                        ("Total Equity", _bs_diff_rs.get("Total Equity"), bs_tn),
                        ("Assets = Liabilities + Equity", _bs_diff_rs.get("Balance Check"), bs_tn),
                    ]
                    _SRD = _Ff(name="Arial", size=10, color="CC0000")
                    for chk_lbl, chk_r, chk_tab in _checks:
                        ws_s.cell(_sr, 2, chk_lbl).font = _SF
                        if chk_r:
                            ws_s.cell(_sr, 4, f"='{chk_tab}'!C{chk_r}").number_format = _NUM
                            ws_s.cell(_sr, 5,
                                f'=IF(ABS(D{_sr})<0.01,"Passed","\u26a0 Needs Review")')
                            ws_s.cell(_sr, 5).font = _SGR
                        else:
                            ws_s.cell(_sr, 4, 0).number_format = _NUM
                            ws_s.cell(_sr, 5, "Passed").font = _SGR
                        _sr += 1
                    _sr += 1

                # Reports Included section
                _sec_bar(_sr, "Reports Included")
                _sr += 1

                _display_names = {
                    "P&L": "Profit & Loss", "Balance Sheet": "Balance Sheet",
                    "AR Aging": "AR Aging", "AP Aging": "AP Aging",
                    "IS GL Summary": "IS GL Summary", "BS GL Summary": "BS GL Summary",
                    "BS Balances": "BS Balances", "IS GL Detail": "IS GL Detail",
                    "BS GL Detail": "BS GL Detail",
                    "GL Summary Validation": "GL Summary Validation",
                    "Portal_IS_Flat": "Portal Income Statement",
                    "Portal_BS_Flat": "Portal Balance Sheet",
                }
                _ref_tabs = ["Mapping Reference"] if "Mapping Reference" in wb_fin.sheetnames else []
                _sec_labels = ["QBO Reports", "Mapped Reports", "Mapping Reference", "Flat Files", "Validation", "Portal Data"]
                _sec_lists  = [qbo_tabs, mapped_tabs, _ref_tabs, flat_tabs, val_tabs, portal_tabs]
                for si, (sl, st) in enumerate(zip(_sec_labels, _sec_lists)):
                    if not st:
                        continue
                    ws_s.cell(_sr, 2, sl).font = _SFB
                    _sr += 1
                    for tn in st:
                        dn = _display_names.get(tn, tn)
                        c = ws_s.cell(_sr, 2, f"\u2192 {dn}")
                        c.font = _SLK
                        c.hyperlink = f"#'{tn}'!A1"
                        _sr += 1
                    _sr += 1

                ws_s.column_dimensions["B"].width = 40
                ws_s.column_dimensions["C"].width = 15
                ws_s.column_dimensions["D"].width = 15
                ws_s.column_dimensions["E"].width = 20

                # ── Reorder tabs ──
                # Build final order, then move each sheet to end in sequence
                _final = []
                for tn in ordered:
                    if tn in wb_fin.sheetnames:
                        _final.append(tn)
                for sn in wb_fin.sheetnames:
                    if sn not in _final:
                        _final.append(sn)
                for i, tn in enumerate(_final):
                    wb_fin.move_sheet(tn, offset=i - wb_fin.sheetnames.index(tn))

                # Arial 10 enforcement on ALL tabs (final pass before save)
                for _ws in wb_fin.worksheets:
                    for _row in _ws.iter_rows(min_row=1, max_row=max(_ws.max_row, 1), max_col=max(_ws.max_column, 1)):
                        for _c in _row:
                            if _c.font and (_c.font.name != "Arial" or _c.font.size != 10):
                                if _c.font.size and _c.font.size > 10:
                                    continue
                                _c.font = _Ff(
                                    name="Arial", size=10,
                                    bold=_c.font.bold, italic=_c.font.italic,
                                    color=_c.font.color, underline=_c.font.underline,
                                    strikethrough=_c.font.strikethrough,
                                )

                wb_fin.save(file_path)
                progress_fn("  Workbook restructured.")
            except Exception as _rs_err:
                import traceback
                logger.warning(f"Workbook restructuring failed: {_rs_err}\n{traceback.format_exc()}")
                progress_fn(f"  WARNING: Restructuring failed — {_rs_err}")

            # Upload to Supabase storage
            storage_path = f"{user_id}/{job_id}/{file_name}"
            with open(file_path, "rb") as f:
                supabase.storage.from_("reports").upload(
                    storage_path,
                    f.read(),
                    {"content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
                )
            url_result = supabase.storage.from_("reports").create_signed_url(
                storage_path, 3600
            )
            file_url = (url_result.get("signedURL")
                        or url_result.get("signedUrl")
                        or url_result.get("signed_url", ""))
            logger.info(f"Signed URL keys: {list(url_result.keys())}, url length: {len(file_url)}")
            update_job(job_id, status="complete", file_url=file_url)

    except Exception as e:
        update_job(job_id, status="failed", error=str(e))
    finally:
        try:
            logger.info("Clearing override tokens in finally block")
            qbo_client.set_override_tokens(None)
        except Exception:
            pass


@router.post("/generate")
def generate_report(body: GenerateRequest, user=Depends(get_current_user)):
    """Kick off a report generation job."""
    s, e = _parse_report_dates(body.start_date, body.end_date)

    plan = (user.app_metadata or {}).get("plan", "basic")

    # Date range limit for basic plan (92 days ≈ 3 months)
    if plan == "basic" and (e - s).days > 92:
        raise HTTPException(
            status_code=403,
            detail="Basic plan is limited to a 3-month date range. Upgrade to Pro for unlimited date ranges.",
        )

    # Feature gates by plan
    if plan == "basic":
        body.dimension = "none"
        body.selected_maps = []
        body.include_gl_detail = False
        body.include_portal_data = False
    elif plan in ("pro", "plus"):
        body.include_portal_data = False
    # admin: no restrictions

    job = create_job(
        user_id=str(user.id),
        realm_id=body.realm_id,
        start_date=body.start_date,
        end_date=body.end_date,
        dimension=body.dimension,
    )
    thread = threading.Thread(
        target=run_report_job,
        args=(job["id"], str(user.id), body.realm_id,
              body.start_date, body.end_date, body.dimension,
              body.selected_maps, body.include_gl_detail,
              body.include_portal_data,
              body.include_ar_aging, body.include_ap_aging),
        daemon=True,
    )
    thread.start()
    return {"job_id": job["id"], "status": "pending"}


@router.get("/job/{job_id}")
def get_job_status(job_id: str, user=Depends(get_current_user)):
    """Poll for job status."""
    job = get_job(job_id)
    if not job or job["user_id"] != str(user.id):
        raise HTTPException(status_code=404, detail="Job not found")
    return job


@router.post("/job/{job_id}/cancel")
def cancel_job(job_id: str, user=Depends(get_current_user)):
    """Mark a job as cancelled."""
    job = get_job(job_id)
    if not job or job["user_id"] != str(user.id):
        raise HTTPException(status_code=404, detail="Job not found")
    update_job(job_id, status="failed", error="Cancelled by user")
    return {"cancelled": True}


@router.get("/download/{job_id}")
def download_report(job_id: str, user=Depends(get_current_user)):
    """Download a completed report file through the backend."""
    from fastapi.responses import Response
    job = get_job(job_id)
    if not job or job["user_id"] != str(user.id):
        raise HTTPException(status_code=404, detail="Job not found")
    if job.get("status") != "complete":
        raise HTTPException(status_code=400, detail="Report not ready")

    supabase = get_supabase()
    user_id = str(user.id)
    try:
        files = supabase.storage.from_("reports").list(f"{user_id}/{job_id}")
        if not files:
            raise HTTPException(status_code=404, detail="Report file not found in storage")
        file_name = files[0]["name"]
        storage_path = f"{user_id}/{job_id}/{file_name}"
        data = supabase.storage.from_("reports").download(storage_path)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Download error: {e}")
        raise HTTPException(status_code=500, detail="Could not retrieve report file")

    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{file_name}"'},
    )


@router.get("/history")
def job_history(user=Depends(get_current_user)):
    """Get recent jobs for the current user."""
    return {"jobs": get_user_jobs(str(user.id))}


