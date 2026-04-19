"""
excel_formatter.py

Shared Excel formatting constants + the single global formatting pass that
runs once per workbook before save. Factored out of reports.py to avoid the
three near-identical "Arial 10 / buffer column / freeze panes" loops that
used to live inside run_report_job.
"""

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ── Shared brand styles ──────────────────────────────────────────────────────

FONT_NAME        = "Arial"
FONT_SIZE        = 10
BUFFER_COL_WIDTH = 0.63              # narrow spacer column A on formatted tabs
FREEZE_ROW       = "A6"              # 5 header rows + 1 data start
NUM_FMT          = '#,##0.00_);(#,##0.00);"-"??;@'

# Brand colours
COLOR_TEAL       = "337E8D"   # header fill
COLOR_ACCENT     = "C97D60"   # warm accent
COLOR_DARK_TEAL  = "07393C"
COLOR_LINK_GREEN = "276221"
COLOR_MUTED      = "5A6B6D"


# Font factories — all anchor on Arial 10 so the workbook has a consistent
# default; individual callers can still build variants (italic, coloured)
# without diverging from the brand.

def plain():
    return Font(name=FONT_NAME, size=FONT_SIZE)


def bold():
    return Font(name=FONT_NAME, size=FONT_SIZE, bold=True)


def header():
    return Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color="FFFFFF")


def header_fill():
    return PatternFill("solid", fgColor=COLOR_TEAL)


# ── Tabs we treat as report tabs (get buffer col + freeze panes) ─────────────
# Dividers and validation tabs are formatted separately by their builders.

_BUFFER_COL_A_TABS = {
    "Summary",
    "P&L", "Balance Sheet", "AR Aging", "AP Aging",
    "IS GL Summary", "BS GL Summary", "IS GL Detail", "BS GL Detail",
    "BS Balances",
    "GL Summary Validation",
    "Mapping Reference",
    "Portal_IS_Flat", "Portal_BS_Flat",
}


def _should_get_buffer_col(sheet_name: str) -> bool:
    if sheet_name in _BUFFER_COL_A_TABS:
        return True
    # Mapped reports: "{map_name} P&L", "{map_name} BS", "{map_name} Validation"
    return (
        sheet_name.endswith(" P&L")
        or sheet_name.endswith(" BS")
        or sheet_name.endswith(" Validation")
    )


def apply_global_formatting(wb) -> None:
    """Single pass over the whole workbook before save.

    - Force every cell's font family to Arial and size to 10 unless the cell
      is deliberately larger (e.g. the big company title on the Summary tab).
    - Hide gridlines on every sheet.
    - Apply the narrow buffer column A + freeze panes on report tabs that
      use the standard 5-row header layout.

    Idempotent — safe to call repeatedly.
    """
    if "Normal" in wb.style_names:
        wb._named_styles["Normal"].font = Font(name=FONT_NAME, size=FONT_SIZE)

    for ws in wb.worksheets:
        # Hide gridlines across the board.
        try:
            ws.sheet_view.showGridLines = False
        except Exception:
            pass

        # Arial-10 enforcement. Respect larger title fonts (anything >10 stays).
        for row in ws.iter_rows(min_row=1,
                                max_row=max(ws.max_row, 1),
                                max_col=max(ws.max_column, 1)):
            for cell in row:
                f = cell.font
                if not f:
                    continue
                if f.size and f.size > FONT_SIZE:
                    continue  # keep big title fonts
                if f.name == FONT_NAME and f.size == FONT_SIZE:
                    continue  # already correct
                cell.font = Font(
                    name=FONT_NAME, size=FONT_SIZE,
                    bold=f.bold, italic=f.italic,
                    color=f.color, underline=f.underline,
                    strikethrough=f.strikethrough,
                )

        # Standard report-tab layout: narrow buffer col A + freeze under header.
        if _should_get_buffer_col(ws.title):
            ws.column_dimensions["A"].width = BUFFER_COL_WIDTH
            try:
                ws.freeze_panes = FREEZE_ROW
            except Exception:
                pass


__all__ = [
    "FONT_NAME", "FONT_SIZE", "BUFFER_COL_WIDTH", "FREEZE_ROW", "NUM_FMT",
    "COLOR_TEAL", "COLOR_ACCENT", "COLOR_DARK_TEAL", "COLOR_LINK_GREEN",
    "COLOR_MUTED",
    "plain", "bold", "header", "header_fill",
    "apply_global_formatting",
]
