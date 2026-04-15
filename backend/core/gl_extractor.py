"""
gl_extractor.py
Acorn Lite — raw GL extraction to Excel (openpyxl only, no win32com).

Three pulls:
  1. IS GL Detail — Income/Expense accounts, two rows per txn (Class + Location unpivot)
  2. BS GL Detail — Asset/Liability/Equity accounts, same unpivot
  3. BS Balances  — one row per account per month-end from BalanceSheet API
"""

import calendar
import json
import math
from datetime import date, datetime
from pathlib import Path
from typing import Callable

import openpyxl
import pandas as pd

from qbo_client import fetch_report, fetch_accounts
from report_parser import parse_general_ledger, parse_financial_statement


# ── Account type sets ────────────────────────────────────────────────────────

class LiteCancelled(Exception):
    """Raised when the user clicks Stop during Acorn Lite extraction."""

# Module-level cache for retry-save after PermissionError
_pending_save: dict = {}


def retry_save() -> dict:
    """Retry saving the last built workbook without re-fetching data.

    Call this after the user closes the file in Excel.
    Returns {"path": str} on success, raises PermissionError if still open.
    """
    if not _pending_save.get("wb") or not _pending_save.get("path"):
        raise RuntimeError("No pending save — run the report first.")
    wb        = _pending_save["wb"]
    save_path = _pending_save["path"]
    try:
        wb.save(save_path)
    except PermissionError:
        raise PermissionError(
            f"Still cannot save to '{save_path.name}' — close it in Excel and try again."
        )
    _pending_save.clear()
    return {"path": str(save_path)}


def _check_cancel(cancel_fn):
    if cancel_fn and cancel_fn():
        raise LiteCancelled("Cancelled by user")


_IS_TYPES = {
    "Income", "Cost of Goods Sold", "Expense", "Other Income", "Other Expense",
}
_BS_TYPES = {
    "Bank", "Accounts Receivable", "Other Current Asset", "Fixed Asset", "Other Asset",
    "Accounts Payable", "Credit Card", "Other Current Liability", "Long Term Liability",
    "Equity",
}


def safe_float(val, default=0.0):
    """Convert QBO value to float safely."""
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return default if math.isnan(val) else float(val)
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return default
        try:
            return float(val.replace(",", ""))
        except ValueError:
            return default
    return default


# ── Helpers ──────────────────────────────────────────────────────────────────

def _month_ends(start_date: str, end_date: str) -> list[date]:
    """Return list of month-end dates in [start_date, end_date]."""
    s = datetime.strptime(start_date, "%Y-%m-%d").date()
    e = datetime.strptime(end_date, "%Y-%m-%d").date()
    result = []
    y, m = s.year, s.month
    while date(y, m, 1) <= e:
        last = calendar.monthrange(y, m)[1]
        result.append(date(y, m, last))
        m += 1
        if m > 12:
            m = 1
            y += 1
    return result


def _month_starts(start_date: str, end_date: str) -> list[date]:
    """Return list of month-start dates spanning [start_date, end_date]."""
    s = datetime.strptime(start_date, "%Y-%m-%d").date()
    e = datetime.strptime(end_date, "%Y-%m-%d").date()
    result = []
    y, m = s.year, s.month
    while date(y, m, 1) <= e:
        result.append(date(y, m, 1))
        m += 1
        if m > 12:
            m = 1
            y += 1
    return result


def _parse_date(val):
    """Parse a date string to a date object."""
    if isinstance(val, date):
        return val
    if isinstance(val, str) and val:
        for fmt in ("%Y-%m-%d", "%m/%d/%Y"):
            try:
                return datetime.strptime(val.split("T")[0], fmt).date()
            except ValueError:
                pass
    return val


def _build_coa_lookup(alias: str, progress_fn: Callable = print) -> dict:
    """Build {account_id: {type, subtype}} from Chart of Accounts.

    Also builds a name-based lookup (stripping account number prefix)
    as fallback when Account_ID doesn't match.
    Returns a single dict with both ID-keyed and name-keyed entries.
    Name keys are prefixed with 'name:' to avoid collision with numeric IDs.
    """
    import re
    lookup = {}
    try:
        accts = fetch_accounts(alias)
        type_counts: dict[str, int] = {}
        for a in (accts or []):
            acct_num = a.get("AcctNum", "").strip()
            name     = a.get("Name", "").strip()
            info = {
                "type":     a.get("AccountType", ""),
                "subtype":  a.get("AccountSubType", ""),
                "acct_num": acct_num,
                "name":     name,
            }
            # Key by ID
            lookup[str(a.get("Id", ""))] = info
            # Key by name (lowercase) for fallback matching
            if name:
                lookup[f"name:{name.lower()}"] = info
                # Also store bare name (strip leading number prefix)
                import re as _re2
                bare = _re2.sub(r'^\d[\d.\-]*\s+', '', name).strip()
                if bare and bare.lower() != name.lower():
                    lookup[f"name:{bare.lower()}"] = info

            # Store by account number
            if acct_num:
                lookup[f"acctnum:{acct_num}"] = info
            # Count types
            t = info["type"]
            type_counts[t] = type_counts.get(t, 0) + 1
        progress_fn(f"  COA account types: {type_counts}")
    except Exception as e:
        progress_fn(f"  WARNING: Could not fetch Chart of Accounts ({e}). Account Type filtering will not work — IS GL and BS GL tabs may be empty.")
    return lookup


# ── GL fetch + unpivot ───────────────────────────────────────────────────────

_GL_EXTRA_COLS = "tx_date,txn_type,doc_num,memo,account_num,klass_name,dept_name,subt_nat_amount,rbal_nat_amount"
_MAX_CELLS = 300_000  # QBO limit is 400k; leave buffer
_CELL_LIMIT_MARKER = "Unable to display more data"


def _fetch_gl_single(alias: str, start_date: str, end_date: str) -> pd.DataFrame:
    """Fetch one GL chunk (default + custom columns merged). No v2 handling.
    Retries on network/timeout errors with exponential backoff."""
    import time as _time
    MAX_RETRIES = 3

    for attempt in range(MAX_RETRIES):
        try:
            raw = fetch_report(alias, "GeneralLedger", {
                "start_date": start_date,
                "end_date": end_date,
                "accounting_method": "Accrual",
            })
            _raw_str = str(raw)
            if _CELL_LIMIT_MARKER in _raw_str:
                raise _CellLimitError(f"QBO cell limit hit for {start_date} to {end_date}")

            df = parse_general_ledger(raw)

            # Merge custom columns (Class, Account#)
            try:
                raw_extra = fetch_report(alias, "GeneralLedger", {
                    "start_date": start_date,
                    "end_date": end_date,
                    "accounting_method": "Accrual",
                    "columns": _GL_EXTRA_COLS,
                })
                df_extra = parse_general_ledger(raw_extra)
                if len(df_extra) == len(df):
                    for c in df_extra.columns:
                        if c in ("Account", "Txn_ID"):
                            continue
                        if c == "Account_ID":
                            if "Account_ID" not in df.columns:
                                df["Account_ID"] = df_extra["Account_ID"].values
                            else:
                                mask = df["Account_ID"].fillna("").astype(str).str.strip().isin(["", "0", "nan"])
                                df.loc[mask, "Account_ID"] = df_extra.loc[mask, "Account_ID"].values
                        elif c not in df.columns:
                            df[c] = df_extra[c].values
            except _CellLimitError:
                pass
            except Exception:
                pass

            return df

        except _CellLimitError:
            raise  # don't retry cell limit errors
        except Exception:
            if attempt < MAX_RETRIES - 1:
                _time.sleep(2 ** attempt)
            else:
                raise

    raise RuntimeError("unreachable")


class _CellLimitError(Exception):
    pass


def _fetch_gl(alias: str, start_date: str, end_date: str,
              progress_fn: Callable = print) -> pd.DataFrame:
    """Fetch GL in auto-sized chunks to stay under QBO's 400k cell limit.

    Always uses v1 API — v2 GL drops Amount/Balance columns entirely.
    """
    from qbo_client import _v2_test_mode, set_v2_test_mode
    _was_v2 = _v2_test_mode
    if _was_v2:
        set_v2_test_mode(False)
        progress_fn("  (v2 test mode suspended for GL fetch)")

    try:
        return _fetch_gl_chunked(alias, start_date, end_date, progress_fn)
    finally:
        if _was_v2:
            set_v2_test_mode(True)


def _fetch_gl_chunked(alias: str, start_date: str, end_date: str,
                      progress_fn: Callable = print) -> pd.DataFrame:
    """Auto-chunk GL fetch based on estimated cell count."""
    months = _month_starts(start_date, end_date)
    num_months = len(months)
    if num_months == 0:
        return pd.DataFrame()

    # Step 1: probe up to 3 months to estimate density — use the densest month
    probe_months_to_check = months[:min(3, num_months)]
    probe_rows = 0
    probe_df = pd.DataFrame()

    for probe_month in probe_months_to_check:
        pm_start = probe_month.isoformat()
        pm_end   = date(probe_month.year, probe_month.month,
                        calendar.monthrange(probe_month.year, probe_month.month)[1]).isoformat()
        try:
            pm_df = _fetch_gl_single(alias, pm_start, pm_end)
            if len(pm_df) > probe_rows:
                probe_rows = len(pm_df)
                probe_df   = pm_df
        except _CellLimitError:
            progress_fn("  WARNING: Single month exceeds QBO cell limit — data may be truncated")
        except Exception:
            break  # network error on probe — proceed with what we have
    num_cols = max(len(probe_df.columns), 12) if not probe_df.empty else 12

    if probe_rows == 0:
        # No data in first month — try full range as single call
        progress_fn("  GL probe: 0 rows in first month, trying full range...")
        try:
            return _fetch_gl_single(alias, start_date, end_date)
        except _CellLimitError:
            progress_fn("  WARNING: Full range exceeds QBO cell limit")
            return pd.DataFrame()

    # Step 2: calculate chunk size
    cells_per_month = int(probe_rows * num_cols * 1.5)  # 1.5x safety buffer
    if cells_per_month <= 0:
        cells_per_month = 1
    months_per_chunk = max(1, min(6, _MAX_CELLS // cells_per_month))

    progress_fn(f"  GL estimate: ~{probe_rows} rows/month, ~{cells_per_month} cells/month, "
                f"pulling in {months_per_chunk}-month chunks ({num_months} months total)")

    # If full range fits in one chunk, just return the single call
    if num_months <= months_per_chunk:
        if num_months == 1:
            # Already have first month
            progress_fn(f"  GL: {probe_rows} rows (single month)")
            return probe_df
        progress_fn(f"  GL: fetching full range as single call...")
        try:
            return _fetch_gl_single(alias, start_date, end_date)
        except _CellLimitError:
            progress_fn("  WARNING: Full range hit cell limit, falling back to chunked fetch")
            months_per_chunk = max(1, months_per_chunk // 2)

    # Step 3: chunked fetch
    all_dfs: list[pd.DataFrame] = []
    _had_retry = False
    chunk_idx = 0
    i = 0
    while i < num_months:
        chunk_end_idx = min(i + months_per_chunk, num_months) - 1
        c_start = months[i].isoformat()
        c_end_month = months[chunk_end_idx]
        c_end = date(c_end_month.year, c_end_month.month,
                     calendar.monthrange(c_end_month.year, c_end_month.month)[1]).isoformat()

        chunk_idx += 1
        n_chunks_est = math.ceil(num_months / months_per_chunk)
        progress_fn(f"  GL chunk [{chunk_idx}/{n_chunks_est}]: {c_start} to {c_end}")

        try:
            chunk_df = _fetch_gl_single(alias, c_start, c_end)
            progress_fn(f"    {len(chunk_df)} rows")
            all_dfs.append(chunk_df)
            i += months_per_chunk
        except _CellLimitError:
            # Halve chunk size and retry this chunk
            _had_retry = True
            old_size = months_per_chunk
            months_per_chunk = max(1, months_per_chunk // 2)
            progress_fn(f"    Cell limit hit — reducing chunk from {old_size} to {months_per_chunk} months")
            if months_per_chunk == old_size:
                progress_fn(f"    ERROR: {c_start} exceeds QBO cell limit even at 1-month chunk — this month will be missing from output. Consider a shorter date range.")
                i += 1

    if not all_dfs:
        return pd.DataFrame()

    # Step 4: combine
    combined = pd.concat(all_dfs, ignore_index=True)

    # Only remove beginning-balance rows (empty Transaction Type) that repeat
    # across chunk boundaries. Do NOT drop_duplicates on real transactions —
    # recurring charges with identical date/amount/description are legitimate.
    before = len(combined)
    if "Transaction Type" in combined.columns:
        real_txns = combined[combined["Transaction Type"].fillna("").str.strip() != ""]
        bbal_rows = combined[combined["Transaction Type"].fillna("").str.strip() == ""]
        bbal_deduped = bbal_rows.drop_duplicates(keep="first")
        combined = pd.concat([real_txns, bbal_deduped], ignore_index=True)
    else:
        combined = combined.drop_duplicates(keep="first")

    dupes = before - len(combined)
    if dupes:
        progress_fn(f"  GL deduplicated: {dupes} beginning-balance rows removed")

    combined = combined.sort_values(
        [c for c in ["Date", "Account"] if c in combined.columns],
        kind="stable"
    ).reset_index(drop=True)

    total_cells = len(combined) * len(combined.columns)
    progress_fn(f"  GL total: {len(combined)} rows, {chunk_idx} chunk(s), ~{total_cells:,} cells")
    if _had_retry:
        progress_fn("  WARNING: Some chunks were split due to cell limit — consider smaller date ranges")

    return combined


# Column rename map (same as generate_report._GL_COL_MAP subset)
_COL_MAP = {
    "date": "Date", "transaction type": "Transaction Type", "num": "Num",
    "no.": "Num", "name": "Customer/Vendor", "memo/description": "Memo/Description",
    "account #": "Account Number", "account no.": "Account Number",
    "account number": "Account Number", "account_num": "Account Number",
    "account type": "Account Type", "split": "Split Account",
    "class": "Class", "klass_name": "Class", "class_name": "Class",
    "department": "Location", "location": "Location", "dept_name": "Location",
    "amount": "Amount", "balance": "Running Balance",
    "subt_nat_amount": "Amount", "rbal_nat_amount": "Running Balance",
}


def _rename_cols(df: pd.DataFrame) -> pd.DataFrame:
    """Case-insensitive column rename."""
    lower_map = {k.lower(): v for k, v in _COL_MAP.items()}
    df.columns = [lower_map.get(c.lower(), c) for c in df.columns]
    return df


def _prepare_gl_df_common(
    df: pd.DataFrame,
    coa_lookup: dict,
    account_types: set,
) -> pd.DataFrame:
    """Filter GL to account types, enrich with COA data, parse dates/amounts."""
    df = _rename_cols(df.copy())

    if "Account" in df.columns:
        df = df.rename(columns={"Account": "Account Name"})

    # Strip beginning-balance rows
    if "Transaction Type" in df.columns:
        has_type = df["Transaction Type"].fillna("").str.strip() != ""
        df = df[has_type].copy()

    # Ensure all columns exist
    for col in ["Account Name", "Account_ID", "Date", "Transaction Type", "Num",
                "Account Number", "Account Type", "Account Subtype",
                "Class", "Location", "Customer/Vendor", "Memo/Description",
                "Split Account", "Amount", "Running Balance"]:
        if col not in df.columns:
            df[col] = ""

    # Fix rows where Account Name is blank but Account Number is populated
    # QBO sometimes returns rows with blank Account but the number is in Account #
    import re as _re_pre
    if "Account Number" in df.columns and "Account Name" in df.columns:
        blank_name_mask = df["Account Name"].fillna("").str.strip() == ""
        has_acct_num_mask = df["Account Number"].fillna("").astype(str).str.strip().isin(["", "0", "nan"]) == False
        fix_mask = blank_name_mask & has_acct_num_mask
        if fix_mask.any():
            for idx in df[fix_mask].index:
                acct_num_val = str(df.at[idx, "Account Number"]).strip().replace(".0", "").strip()
                acct_num_info = coa_lookup.get(f"acctnum:{acct_num_val}")
                if acct_num_info:
                    df.at[idx, "Account Name"] = f"{acct_num_val} {acct_num_info.get('name', '')}".strip()
                    df.at[idx, "Account Type"] = acct_num_info.get("type", "")
                    df.at[idx, "Account Subtype"] = acct_num_info.get("subtype", "")
            fixed = fix_mask.sum()
            if fixed:
                print(f"  Fixed {fixed} rows with blank Account Name using Account Number lookup")

    # Enrich Account Type / Subtype from COA
    # Try Account_ID first, then fall back to account name (strip number prefix)
    import re as _re
    _id_hits = 0
    _name_hits = 0
    _misses = 0
    if coa_lookup:
        for idx, row in df.iterrows():
            if str(row.get("Account Type", "")).strip():
                continue  # already has a type
            acct_id = str(row.get("Account_ID", "")).strip()
            if acct_id.endswith(".0"):
                acct_id = acct_id[:-2]
            info = coa_lookup.get(acct_id) if acct_id else None
            if not info:
                acct_name = str(row.get("Account Name", "")).strip()
                bare_name = _re.sub(r'^\d[\d.\-]*\s+', '', acct_name).strip()
                info = coa_lookup.get(f"name:{bare_name.lower()}")
                if not info and acct_name:
                    info = coa_lookup.get(f"name:{acct_name.lower()}")

                # Final fallback: look up by Account Number column
                if not info:
                    acct_num_val = str(row.get("Account Number", "") or "").strip()
                    if not acct_num_val:
                        # Try extracting number prefix from Account Name
                        import re as _re2
                        num_match = _re2.match(r'^(\d[\d.\-]*)', acct_name or "")
                        if num_match:
                            acct_num_val = num_match.group(1)
                    if acct_num_val:
                        acct_num_info = coa_lookup.get(f"acctnum:{acct_num_val}")
                        if acct_num_info:
                            info = acct_num_info
                            # If Account Name is blank, fill it from the COA
                            if not acct_name and acct_num_info.get("name"):
                                df.at[idx, "Account Name"] = f"{acct_num_val} {acct_num_info['name']}"

                if info:
                    _name_hits += 1
                else:
                    _misses += 1
                    continue
            else:
                _id_hits += 1
            df.at[idx, "Account Type"] = info["type"]
            df.at[idx, "Account Subtype"] = info["subtype"]

    df["Account Type"] = df["Account Type"].fillna("").astype(str).str.strip()
    _type_counts = df["Account Type"].value_counts().to_dict()
    _is_total = sum(v for k, v in _type_counts.items() if k in _IS_TYPES)
    _bs_total = sum(v for k, v in _type_counts.items() if k in _BS_TYPES)
    print(f"  COA join: {_id_hits} by ID, {_name_hits} by name, {_misses} unmatched — IS: {_is_total}, BS: {_bs_total}")

    # Filter to requested account types
    df = df[df["Account Type"].isin(account_types)].copy()

    if df.empty:
        return df

    df["Date"] = df["Date"].apply(_parse_date)
    df["Amount"] = df["Amount"].apply(safe_float)
    df["Running Balance"] = df["Running Balance"].apply(safe_float)
    df = df.sort_values(["Date", "Account Name"], kind="stable").reset_index(drop=True)
    return df


def _prepare_is_gl_rows(
    df: pd.DataFrame,
    coa_lookup: dict,
    progress_fn: Callable = print,
    dimension: str = "class",
) -> list[list]:
    """IS GL Detail — rows per transaction with dimension handling.

    dimension:
      "none"     — one row per txn, Class and Location as separate columns
      "class"    — one row per txn, Class as last column
      "location" — one row per txn, Location as last column
    """
    df = _prepare_gl_df_common(df, coa_lookup, _IS_TYPES)
    if df.empty:
        return []

    dimension = (dimension or "class").lower().strip()

    if dimension == "none":
        header = [
            "Date", "Month", "Transaction Type", "Num", "Account Name", "Account Number",
            "Account Type", "Account Subtype", "Class", "Location",
            "Customer/Vendor", "Memo/Description", "Split Account",
            "Amount", "Running Balance",
        ]
        rows = [header]
        for _, r in df.iterrows():
            d = r["Date"]
            eom = date(d.year, d.month, calendar.monthrange(d.year, d.month)[1]) if isinstance(d, date) else d
            rows.append([
                r["Date"], eom, r["Transaction Type"], r.get("Num", ""),
                r["Account Name"], r.get("Account Number", ""),
                r["Account Type"], r.get("Account Subtype", ""),
                str(r.get("Class", "") or ""), str(r.get("Location", "") or ""),
                r.get("Customer/Vendor", ""), r.get("Memo/Description", ""),
                r.get("Split Account", ""),
                r["Amount"], r["Running Balance"],
            ])
        return rows

    elif dimension == "location":
        header = [
            "Date", "Month", "Transaction Type", "Num", "Account Name", "Account Number",
            "Account Type", "Account Subtype",
            "Customer/Vendor", "Memo/Description", "Split Account",
            "Amount", "Running Balance",
            "Location",
        ]
        rows = [header]
        for _, r in df.iterrows():
            d = r["Date"]
            eom = date(d.year, d.month, calendar.monthrange(d.year, d.month)[1]) if isinstance(d, date) else d
            rows.append([
                r["Date"], eom, r["Transaction Type"], r.get("Num", ""),
                r["Account Name"], r.get("Account Number", ""),
                r["Account Type"], r.get("Account Subtype", ""),
                r.get("Customer/Vendor", ""), r.get("Memo/Description", ""),
                r.get("Split Account", ""),
                r["Amount"], r["Running Balance"],
                str(r.get("Location", "") or ""),
            ])
        return rows

    else:
        # Default: "class"
        header = [
            "Date", "Month", "Transaction Type", "Num", "Account Name", "Account Number",
            "Account Type", "Account Subtype",
            "Customer/Vendor", "Memo/Description", "Split Account",
            "Amount", "Running Balance",
            "Class",
        ]
        rows = [header]
        for _, r in df.iterrows():
            d = r["Date"]
            eom = date(d.year, d.month, calendar.monthrange(d.year, d.month)[1]) if isinstance(d, date) else d
            rows.append([
                r["Date"], eom, r["Transaction Type"], r.get("Num", ""),
                r["Account Name"], r.get("Account Number", ""),
                r["Account Type"], r.get("Account Subtype", ""),
                r.get("Customer/Vendor", ""), r.get("Memo/Description", ""),
                r.get("Split Account", ""),
                r["Amount"], r["Running Balance"],
                str(r.get("Class", "") or ""),
            ])
        return rows


def _prepare_bs_gl_rows(
    df: pd.DataFrame,
    coa_lookup: dict,
    progress_fn: Callable = print,
) -> list[list]:
    """BS GL Detail — one row per transaction, no Class/Location unpivot."""
    df = _prepare_gl_df_common(df, coa_lookup, _BS_TYPES)
    if df.empty:
        return []

    header = [
        "Date", "Month", "Transaction Type", "Num", "Account Name", "Account Number",
        "Account Type", "Account Subtype",
        "Customer/Vendor", "Memo/Description", "Split Account",
        "Amount", "Running Balance",
    ]
    rows = [header]

    for _, r in df.iterrows():
        d = r["Date"]
        eom = date(d.year, d.month, calendar.monthrange(d.year, d.month)[1]) if isinstance(d, date) else d
        rows.append([
            r["Date"], eom, r["Transaction Type"], r.get("Num", ""),
            r["Account Name"], r.get("Account Number", ""),
            r["Account Type"], r.get("Account Subtype", ""),
            r.get("Customer/Vendor", ""), r.get("Memo/Description", ""),
            r.get("Split Account", ""),
            r["Amount"], r["Running Balance"],
        ])

    return rows


# ── BS Balances ──────────────────────────────────────────────────────────────

def _fetch_bs_balances(
    alias: str,
    start_date: str,
    end_date: str,
    progress_fn: Callable = print,
    coa_lookup: dict | None = None,
) -> list[list]:
    """Fetch BS one call per month, return flat rows."""
    months = _month_ends(start_date, end_date)
    header = ["Account", "Account Type", "Account Subtype", "Account Group", "Month", "Ending Balance"]
    rows = [header]

    # Map QBO account types to simple groups
    _ASSET_TYPES    = {"Bank", "Accounts Receivable", "Other Current Asset",
                       "Fixed Asset", "Other Asset"}
    _LIAB_TYPES     = {"Accounts Payable", "Credit Card", "Other Current Liability",
                       "Long Term Liability"}
    _EQUITY_TYPES   = {"Equity"}

    def _account_group(acct_type: str) -> str:
        if acct_type in _ASSET_TYPES:
            return "Asset"
        if acct_type in _LIAB_TYPES:
            return "Liability"
        if acct_type in _EQUITY_TYPES:
            return "Equity"
        return ""

    for me_dt in months:
        progress_fn(f"  BS {me_dt.strftime('%b %Y')}...")
        raw = fetch_report(alias, "BalanceSheet", {
            "start_date": start_date,
            "end_date": me_dt.isoformat(),
            "accounting_method": "Accrual",
        })
        df = parse_financial_statement(raw)
        if df.empty:
            continue

        # Get amount column for this single-month report
        amount_col = None
        meta = {"Row_Type", "Indent_Level", "Account_Path", "Account", "Account_ID"}
        for c in df.columns:
            if c not in meta:
                if pd.api.types.is_numeric_dtype(df[c]) or c == "Total":
                    amount_col = c
                    break
        if amount_col is None:
            non_meta = [c for c in df.columns if c not in meta]
            if non_meta:
                amount_col = non_meta[-1]

        if amount_col is None:
            continue

        # Truncate at GrandTotal — stop before it, exclude everything after
        grand_total_idx = df[df["Row_Type"] == "GrandTotal"].index
        if len(grand_total_idx) > 0:
            df = df.loc[:grand_total_idx[0]]
            df = df[df["Row_Type"] != "GrandTotal"]
        data_rows = df[df["Row_Type"].notna()].copy()
        import re as _re_bs
        for _, r in data_rows.iterrows():
            acct = str(r.get("Account", "")).strip()
            if not acct:
                continue
            bal = safe_float(r.get(amount_col))

            # Look up account type from COA
            acct_type = ""
            acct_subtype = ""
            if coa_lookup:
                bare = _re_bs.sub(r'^\d[\d.\-]*\s+', '', acct).strip()
                info = (coa_lookup.get(f"name:{acct.lower()}") or
                        coa_lookup.get(f"name:{bare.lower()}"))
                if info:
                    acct_type    = info.get("type", "")
                    acct_subtype = info.get("subtype", "")

            acct_group = _account_group(acct_type)

            # Net Income is a calculated row QBO inserts into the BS equity section.
            # It has no COA entry so the lookup returns blank — tag it as Equity manually.
            if not acct_group and "net income" in acct.lower():
                acct_group = "Equity"
                acct_type  = acct_type or "Equity"

            rows.append([acct, acct_type, acct_subtype, acct_group, me_dt, bal])

    return rows


# ── Monthly P&L and Balance Sheet ────────────────────────────────────────────

def _fetch_monthly_reports(
    alias: str,
    start_date: str,
    end_date: str,
    progress_fn: Callable = print,
) -> tuple[list[list], list[list]]:
    """Fetch monthly P&L and Balance Sheet data.

    P&L: ONE API call using summarize_column_by=Month — all months in one response.
    BS:  One call per month-end (QBO limitation — BS is always point-in-time).
    """
    months    = _month_ends(start_date, end_date)
    meta_cols = {"Row_Type", "Indent_Level", "Account_Path", "Account", "Account_ID"}

    # ── P&L: single call ──────────────────────────────────────────────────
    pl_rows: list[list] = []
    try:
        progress_fn("  Fetching P&L (single call, all months)...")
        pl_raw = fetch_report(alias, "ProfitAndLoss", {
            "start_date":          start_date,
            "end_date":            end_date,
            "accounting_method":   "Accrual",
            "summarize_column_by": "Month",
        })
        pl_df = parse_financial_statement(pl_raw)

        if not pl_df.empty:
            month_cols = [c for c in pl_df.columns if c not in meta_cols]
            # Remove QBO-added Total column — it's not a month and would break cross-checks
            month_cols = [c for c in month_cols if str(c).strip().lower() not in ("total", "totals")]

            def _fmt_col(col):
                col = str(col).strip()
                try:
                    from datetime import datetime as _dt
                    d = _dt.strptime(col[:7], "%Y-%m")
                    return d.strftime("%b %Y")
                except Exception:
                    return col

            month_labels = [_fmt_col(c) for c in month_cols]
            pl_rows.append(["Account"] + month_labels + ["Total"])

            for _, row in pl_df.iterrows():
                acct      = str(row.get("Account", "") or "").strip()
                indent    = int(row.get("Indent_Level", 0) or 0)
                row_type  = str(row.get("Row_Type", "") or "")
                if not acct:
                    continue
                label  = ("  " * indent) + acct
                values = [safe_float(row.get(c)) for c in month_cols]
                total  = sum(values)
                # Store row_type as last element for lookup use (not written to Excel)
                pl_rows.append([label] + values + [total] + [row_type])

            progress_fn(f"  P&L: {len(pl_rows)-1} rows, {len(month_labels)} months (single call)")
    except Exception as e:
        progress_fn(f"  WARNING: P&L fetch failed — {e}")

    # ── BS: one call per month-end ────────────────────────────────────────
    bs_rows: list[list] = []
    if not months:
        return pl_rows, bs_rows

    bs_data:         dict[str, dict] = {}
    bs_order:        list[str]       = []
    bs_month_labels: list[str]       = [m.strftime("%b %Y") for m in months]

    # Step 1: One call for the FULL period to get complete account list in correct order
    progress_fn(f"  Balance Sheet: fetching account structure...")
    try:
        bs_structure_raw = fetch_report(alias, "BalanceSheet", {
            "start_date":        start_date,
            "end_date":          months[-1].isoformat(),
            "accounting_method": "Accrual",
        })
        bs_structure_df = parse_financial_statement(bs_structure_raw)
        if not bs_structure_df.empty:
            for _, row in bs_structure_df.iterrows():
                row_type = str(row.get("Row_Type", "") or "")
                if row_type == "GrandTotal":
                    break
                acct   = str(row.get("Account", "") or "").strip()
                indent = int(row.get("Indent_Level", 0) or 0)
                if not acct:
                    continue
                if acct not in bs_data:
                    bs_data[acct] = {
                        "indent":   indent,
                        "row_type": row_type,
                        "values":   [0.0] * len(months),
                    }
                    bs_order.append(acct)
    except Exception as e:
        progress_fn(f"  WARNING: BS structure fetch failed \u2014 {e}")

    # Step 2: One call per month-end to get balances — only update known accounts
    for mi, me_dt in enumerate(months):
        progress_fn(f"  Balance Sheet: {me_dt.strftime('%b %Y')}...")
        try:
            bs_raw = fetch_report(alias, "BalanceSheet", {
                "start_date":        start_date,
                "end_date":          me_dt.isoformat(),
                "accounting_method": "Accrual",
            })
            bs_df = parse_financial_statement(bs_raw)
            if bs_df.empty:
                continue
            bs_num_cols = [c for c in bs_df.columns if c not in meta_cols]
            for _, row in bs_df.iterrows():
                row_type = str(row.get("Row_Type", "") or "")
                if row_type == "GrandTotal":
                    break
                acct = str(row.get("Account", "") or "").strip()
                if not acct or acct not in bs_data:
                    continue
                amount = 0.0
                for c in reversed(bs_num_cols):
                    v = safe_float(row.get(c))
                    if v != 0.0:
                        amount = v
                        break
                bs_data[acct]["values"][mi] = amount
        except Exception as e:
            progress_fn(f"  WARNING: BS fetch failed for {me_dt.strftime('%b %Y')} \u2014 {e}")

    # Write BS rows — in the order established by the full-period structure call
    if bs_order:
        bs_rows.append(["Account"] + bs_month_labels)
        for acct in bs_order:
            info  = bs_data[acct]
            label = ("  " * info["indent"]) + acct
            bs_rows.append([label] + info["values"])

    progress_fn(f"  Monthly reports: {len(pl_rows)-1} P&L rows, {len(bs_rows)-1} BS rows")
    return pl_rows, bs_rows


# ── IS by Dimension (Class / Location) ───────────────────────────────────────

def _fetch_pl_by_dimension(
    alias: str,
    start_date: str,
    end_date: str,
    dimension: str,
    progress_fn: Callable = print,
) -> list[list]:
    """Fetch P&L by Class or Location, unpivot into flat rows."""
    import re as _re

    dim_label     = "Class" if dimension == "class" else "Location"
    # QBO accepts "Classes" for class breakdown, "Departments" for location
    qbo_summarize = "Classes" if dimension == "class" else "Departments"
    meta_cols     = {"Row_Type", "Indent_Level", "Account_Path", "Account", "Account_ID"}

    months = _month_ends(start_date, end_date)
    if not months:
        return []

    progress_fn(f"  Fetching IS by {dim_label}...")

    account_data: dict[str, dict] = {}
    account_order: list[str] = []
    all_dims: set[str] = set()
    month_labels = [m.strftime("%b %Y") for m in months]

    for mi, me_dt in enumerate(months):
        month_start = date(me_dt.year, me_dt.month, 1).isoformat()
        month_end   = me_dt.isoformat()
        month_label = month_labels[mi]

        try:
            raw = fetch_report(alias, "ProfitAndLoss", {
                "start_date":          month_start,
                "end_date":            month_end,
                "accounting_method":   "Accrual",
                "summarize_column_by": qbo_summarize,
            })
            df = parse_financial_statement(raw)
            if df.empty:
                progress_fn(f"  DEBUG: parse_financial_statement returned empty for {month_label} {dim_label}")
                continue

            dim_cols = [c for c in df.columns if c not in meta_cols]
            # Remove QBO-added Total/subtotal columns
            dim_cols = [c for c in dim_cols if c.lower() not in ("total", "totals", "")]
            progress_fn(f"  DEBUG IS by {dim_label} {month_label}: {len(df)} rows, dim_cols={dim_cols[:5]}")

            if not dim_cols:
                progress_fn(f"  WARNING: No dimension columns found for {month_label} — QBO may not have returned summarize_column_by data")
                try:
                    cols_raw = raw.get("Columns", {}).get("Column", [])
                    progress_fn(f"  DEBUG raw columns: {[c.get('ColTitle','') for c in cols_raw[:10]]}")
                except Exception:
                    pass
                continue

            all_dims.update(dim_cols)

            for _, row in df.iterrows():
                acct     = str(row.get("Account", "") or "").strip()
                row_type = str(row.get("Row_Type", "") or "")
                indent   = int(row.get("Indent_Level", 0) or 0)
                if not acct:
                    continue

                key = f"{acct}||{row_type}"
                if key not in account_data:
                    account_data[key] = {
                        "account": acct, "row_type": row_type, "indent": indent,
                        "dim_data": {},
                    }
                    account_order.append(key)

                for dim_col in dim_cols:
                    amount = safe_float(row.get(dim_col))
                    if amount == 0.0:
                        continue
                    if dim_col not in account_data[key]["dim_data"]:
                        account_data[key]["dim_data"][dim_col] = {}
                    account_data[key]["dim_data"][dim_col][month_label] = amount

        except Exception as e:
            progress_fn(f"  WARNING: IS by {dim_label} fetch failed for {month_label} — {e}")
            continue

    if not account_order:
        progress_fn(f"  WARNING: No data returned for IS by {dim_label}")
        return []

    # Unpivot into flat rows
    header = ["Account", "Row Type", "Indent Level", dim_label] + month_labels
    rows   = [header]

    sorted_dims = sorted(all_dims, key=lambda d: ("" if d else "~", d.lower() if d else ""))

    for key in account_order:
        info       = account_data[key]
        acct       = info["account"]
        row_type   = info["row_type"]
        indent     = info["indent"]
        active_dims = set(info["dim_data"].keys())

        if not active_dims:
            continue

        for dim_val in sorted_dims:
            if dim_val not in active_dims:
                continue
            monthly_amounts = info["dim_data"][dim_val]
            month_values    = [monthly_amounts.get(ml, 0.0) for ml in month_labels]
            if all(v == 0.0 for v in month_values):
                continue
            rows.append([acct, row_type, indent, dim_val] + month_values)

    progress_fn(f"  IS by {dim_label}: {len(rows)-1} rows")
    return rows


# ── AR / AP Aging ────────────────────────────────────────────────────────────

def _parse_aging_report(report_json: dict) -> list[list]:
    """Parse a QBO AgedReceivables or AgedPayables report into flat rows.

    Returns [header_row, data_row_1, ...] where column names are read
    dynamically from the API response.
    """
    columns = report_json.get("Columns", {}).get("Column", [])
    headers = [c.get("ColTitle", "") for c in columns]
    if not headers:
        return []

    rows_section = report_json.get("Rows", {}).get("Row", [])
    data_rows: list[list] = []

    def _walk_rows(row_list, section_name=""):
        for row in row_list:
            row_type = row.get("type", "")
            col_data = row.get("ColData", [])
            # Header row for a section (customer/vendor group)
            if row_type == "Section":
                header_row = row.get("Header", {})
                sub_rows = row.get("Rows", {}).get("Row", [])
                sec_name = ""
                if header_row and header_row.get("ColData"):
                    sec_name = header_row["ColData"][0].get("value", "")
                _walk_rows(sub_rows, sec_name)
                # Summary (subtotal) row for the section
                summary = row.get("Summary", {})
                if summary and summary.get("ColData"):
                    vals = [cd.get("value", "") for cd in summary["ColData"]]
                    parsed = _parse_row_values(vals)
                    data_rows.append(("subtotal", parsed))
            elif col_data:
                vals = [cd.get("value", "") for cd in col_data]
                parsed = _parse_row_values(vals)
                row_label = vals[0] if vals else ""
                if row_label.lower().startswith("total"):
                    data_rows.append(("total", parsed))
                else:
                    data_rows.append(("data", parsed))

    def _parse_row_values(vals):
        """Convert numeric strings to floats, leave others as-is."""
        out = []
        for i, v in enumerate(vals):
            if i == 0:
                out.append(v)
            else:
                try:
                    out.append(float(v))
                except (ValueError, TypeError):
                    out.append(v if v else "")
        return out

    _walk_rows(rows_section)

    result = [headers]
    for _kind, row_vals in data_rows:
        result.append(row_vals)
    return result


def _fetch_ar_aging(alias, as_of_date, progress_fn):
    """Fetch AR Aging (AgedReceivables) report from QBO."""
    raw = fetch_report(alias, "AgedReceivables", {
        "report_date": as_of_date,
    })
    rows = _parse_aging_report(raw)
    progress_fn(f"  AR Aging: {len(rows) - 1 if rows else 0} rows")
    return rows


def _fetch_ap_aging(alias, as_of_date, progress_fn):
    """Fetch AP Aging (AgedPayables) report from QBO."""
    raw = fetch_report(alias, "AgedPayables", {
        "report_date": as_of_date,
    })
    rows = _parse_aging_report(raw)
    progress_fn(f"  AP Aging: {len(rows) - 1 if rows else 0} rows")
    return rows


def _write_aging_sheet(wb: openpyxl.Workbook, tab_name: str, rows: list[list],
                       end_date: str):
    """Write an AR Aging or AP Aging tab with subtitle, formatted headers, and data."""
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(tab_name)
    ws.sheet_view.showGridLines = False

    if not rows or len(rows) < 2:
        ws.cell(row=1, column=1, value="No data").font = PLAIN_FONT
        return

    # Row 1: subtitle — "As of Month DD, YYYY"
    dt = datetime.strptime(end_date, "%Y-%m-%d")
    subtitle = f"As of {dt.strftime('%B %d, %Y')}"
    ws.cell(row=1, column=1, value=subtitle).font = BOLD_FONT

    # Row 2: column headers
    headers = rows[0]
    for ci, val in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=val)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center" if ci > 1 else "left")

    # Parse the aging data to identify row types for formatting
    parsed = _parse_aging_report.__code__  # we already have parsed rows
    # Re-derive row types from the data
    data_rows = rows[1:]
    last_idx = len(data_rows) - 1

    for ri, row in enumerate(data_rows):
        excel_row = ri + 3  # data starts at row 3
        label = str(row[0] or "").strip().lower() if row else ""
        is_total = label.startswith("total")
        is_grand = (ri == last_idx and is_total)

        for ci in range(1, len(row) + 1):
            val = row[ci - 1] if (ci - 1) < len(row) else None
            c = ws.cell(row=excel_row, column=ci, value=val)
            c.font = BOLD_FONT if is_total else PLAIN_FONT

            if ci == 1:
                c.alignment = Alignment(horizontal="left")
            elif isinstance(val, (int, float)):
                c.number_format = _ACCT_FMT
                c.alignment = Alignment(horizontal="right")

            if is_grand:
                c.border = _Bdr(top=THIN, bottom=DOUBLE)
            elif is_total:
                c.border = _Bdr(top=THIN)

    # Autofit column widths
    for ci in range(1, len(headers) + 1):
        max_len = len(str(headers[ci - 1] or ""))
        for ri in range(len(data_rows)):
            cell_val = data_rows[ri][ci - 1] if (ci - 1) < len(data_rows[ri]) else ""
            max_len = max(max_len, len(str(cell_val or "")))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 3, 40)

    # Freeze below subtitle and header
    ws.freeze_panes = "A3"


# ── Validation (live formula version) ────────────────────────────────────────

def _fetch_qbo_report_totals(
    alias: str,
    start_date: str,
    end_date: str,
    progress_fn: Callable = print,
    coa_lookup: dict | None = None,
) -> tuple[dict, dict]:
    """Fetch QBO P&L and Balance Sheet report totals for validation.

    Returns (qbo_is, qbo_bs) dicts mapping account_name → {"amount": float, "acct_num": str}.
    """
    import re as _re

    def _bare(name: str) -> str:
        return _re.sub(r'^\d[\d.\-]*\s+', '', str(name or "")).strip()

    def _lookup_acct_num(acct_id: str, acct_name: str) -> str:
        """Look up account number from COA lookup by ID or name."""
        if not coa_lookup:
            return ""
        info = coa_lookup.get(str(acct_id).strip()) if acct_id else None
        if not info:
            bare = _bare(acct_name)
            info = (coa_lookup.get(f"name:{acct_name.lower()}") or
                    coa_lookup.get(f"name:{bare.lower()}"))
        return (info or {}).get("acct_num", "") if info else ""

    qbo_is: dict[str, dict] = {}
    try:
        pl_raw = fetch_report(alias, "ProfitAndLoss", {
            "start_date": start_date,
            "end_date":   end_date,
            "accounting_method": "Accrual",
        })
        pl_df  = parse_financial_statement(pl_raw)
        meta   = {"Row_Type", "Indent_Level", "Account_Path", "Account", "Account_ID"}
        all_num_cols = [c for c in pl_df.columns if c not in meta]
        # Use only the first numeric column — QBO with no summarize_column_by
        # returns a single total column. Taking only first avoids double-counting.
        num_cols = [all_num_cols[0]] if all_num_cols else []

        for _, row in pl_df.iterrows():
            if row.get("Row_Type") not in ("Data", "Section"):
                continue
            acct = str(row.get("Account", "") or "").strip()
            if not acct:
                continue
            total    = sum(safe_float(row.get(c)) for c in num_cols)
            acct_id  = str(row.get("Account_ID", "") or "").strip()
            acct_num = _lookup_acct_num(acct_id, acct)
            qbo_is[acct] = {"amount": total, "acct_num": acct_num}
            bare = _bare(acct)
            if bare and bare != acct:
                qbo_is[bare] = {"amount": total, "acct_num": acct_num}
        progress_fn(f"  Validation: {len(qbo_is)} IS account entries from QBO P&L")
    except Exception as e:
        progress_fn(f"  Validation WARNING: Could not fetch P&L — {e}")

    # Remove calculated/subtotal rows that don't exist as real GL accounts
    _is_before = len(qbo_is)
    qbo_is = {
        k: v for k, v in qbo_is.items()
        if k.lower() not in {
            "gross profit", "net income", "net operating income",
            "net other income", "net earnings", "total income",
            "total expenses", "total cost of goods sold",
            "total other income", "total other expense",
        }
        and not k.lower().startswith("total ")
        and not k.lower().startswith("net ")
        and not k.lower().startswith("gross ")
    }
    if _is_before != len(qbo_is):
        progress_fn(f"  Validation: filtered {_is_before - len(qbo_is)} IS calculated rows, {len(qbo_is)} remaining")

    qbo_bs: dict[str, dict] = {}
    try:
        bs_raw = fetch_report(alias, "BalanceSheet", {
            "start_date": start_date,
            "end_date":   end_date,
            "accounting_method": "Accrual",
        })
        bs_df  = parse_financial_statement(bs_raw)
        meta   = {"Row_Type", "Indent_Level", "Account_Path", "Account", "Account_ID"}
        num_cols = [c for c in bs_df.columns if c not in meta]
        for _, row in bs_df.iterrows():
            if row.get("Row_Type") not in ("Data", "Section"):
                continue
            acct = str(row.get("Account", "") or "").strip()
            if not acct:
                continue
            val = 0.0
            for c in reversed(num_cols):
                v = safe_float(row.get(c))
                if v != 0.0:
                    val = v
                    break
            acct_id  = str(row.get("Account_ID", "") or "").strip()
            acct_num = _lookup_acct_num(acct_id, acct)
            qbo_bs[acct] = {"amount": val, "acct_num": acct_num}
            bare = _bare(acct)
            if bare and bare != acct:
                qbo_bs[bare] = {"amount": val, "acct_num": acct_num}
        progress_fn(f"  Validation: {len(qbo_bs)} BS account entries from QBO Balance Sheet")
    except Exception as e:
        progress_fn(f"  Validation WARNING: Could not fetch Balance Sheet — {e}")

    _bs_before = len(qbo_bs)
    qbo_bs = {
        k: v for k, v in qbo_bs.items()
        if k.lower() not in {
            "net income", "retained earnings", "total equity",
            "total liabilities", "total assets",
            "total liabilities and equity",
            "total current assets", "total fixed assets",
            "total other assets", "total current liabilities",
            "total long-term liabilities",
        }
        and not k.lower().startswith("total ")
        and not k.lower().startswith("net ")
    }
    if _bs_before != len(qbo_bs):
        progress_fn(f"  Validation: filtered {_bs_before - len(qbo_bs)} BS calculated rows, {len(qbo_bs)} remaining")

    return qbo_is, qbo_bs


# ── Excel writer ─────────────────────────────────────────────────────────────

_ACCT_NUM_COLS = {"Account Number", "Account #", "Num"}


def _to_date(val):
    """Strip time component from any date/datetime value."""
    if val is None:
        return None
    if hasattr(val, 'date') and callable(val.date):
        return val.date()
    return val

# ── Acorn output formatting constants ────────────────────────────────────────
from openpyxl.styles import Font as _Font, PatternFill as _PF, Alignment as _AL, Border as _Bdr, Side as _Sd

_ARIAL     = "Arial"
_FONT_SZ   = 10
_HDR_COLOR = "337E8D"
_ACCT_FMT  = '#,##0.00_);(#,##0.00);"-"??;@'

def _font(bold=False, color="000000", italic=False):
    return _Font(name=_ARIAL, size=_FONT_SZ, bold=bold, color=color, italic=italic)

HDR_FONT   = _font(bold=True, color="FFFFFF")
HDR_FILL   = _PF("solid", fgColor=_HDR_COLOR)
PLAIN_FONT = _font()
BOLD_FONT  = _font(bold=True)
GRAY_FILL  = _PF("solid", fgColor="F2F2F2")
NO_FILL    = _PF(fill_type=None)
THIN       = _Sd(style="thin")
DOUBLE     = _Sd(style="double")

def _write_sheet(wb: openpyxl.Workbook, tab_name: str, rows: list[list]):
    """Write rows to a worksheet with formatted header, dates, and amounts."""
    from openpyxl.styles import Alignment

    from openpyxl.utils import get_column_letter as _gcl_ws
    ws = wb.create_sheet(tab_name)
    ws.sheet_view.showGridLines = False
    if not rows:
        ws.cell(row=1, column=1, value="No data").font = PLAIN_FONT
        return

    header = rows[0]
    acct_num_cols = {ci for ci, h in enumerate(header) if h in _ACCT_NUM_COLS}

    for ci, col_name in enumerate(header, 1):
        c = ws.cell(row=1, column=ci, value=col_name)
        c.font      = HDR_FONT
        c.fill      = HDR_FILL
        c.alignment = _AL(horizontal="center" if ci > 1 else "left",
                          vertical="center")

    for ri, row in enumerate(rows[1:], 2):
        for ci, val in enumerate(row, 1):
            val = _to_date(val)
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = PLAIN_FONT
            if isinstance(val, date):
                c.number_format = "M/D/YYYY"
            elif isinstance(val, (int, float)):
                c.number_format = "0" if (ci-1) in acct_num_cols else _ACCT_FMT
                c.alignment     = _AL(horizontal="right")

    # Autofit column widths
    for col_cells in ws.columns:
        max_len = 0
        col_letter = _gcl_ws(col_cells[0].column)
        for cell in col_cells:
            try:
                cl = len(str(cell.value)) if cell.value is not None else 0
                if cl > max_len: max_len = cl
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    # Freeze header row
    ws.freeze_panes = "A2"


def _write_validation_sheet(
    wb: openpyxl.Workbook,
    qbo_is: dict,
    qbo_bs: dict,
    start_date: str = "",
    end_date: str = "",
    is_tab_name: str = "IS GL Detail",
    bs_tab_name: str = "BS GL Detail",
    progress_fn: Callable = print,
    dimension: str = "class",
    pl_rows: list | None = None,
    bs_report_rows: list | None = None,
    is_summary_rows: list | None = None,
    bs_summary_rows: list | None = None,
):
    """Write Validation tab with live Excel formulas.

    Column C uses SUMIF/LOOKUP formulas referencing IS/BS GL Detail tabs.
    Column D is static from QBO API. Columns E-F are formulas.
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
    import re as _re

    ws = wb.create_sheet("Validation")
    ws.sheet_view.showGridLines = False

    GREEN      = _PF("solid", fgColor="C6EFCE")
    RED        = _PF("solid", fgColor="FFC7CE")
    YELLOW     = _PF("solid", fgColor="FFEB9C")
    HEADER_BG  = HDR_FILL
    SECTION_BG = _PF("solid", fgColor="BDD7EE")
    BOLD       = BOLD_FONT
    BOLD_WHITE = HDR_FONT
    NUM_FMT    = _ACCT_FMT
    TOLERANCE  = 0.02

    # IS GL Summary column positions — resolve dynamically from summary header
    # Summary format: Month | Account Name | Account Type | (Class/Location) | Amount | map cols...
    from openpyxl.utils import get_column_letter as _gcl_val
    _is_sum_hdr = is_summary_rows[0] if is_summary_rows and len(is_summary_rows) > 0 else []
    _bs_sum_hdr = bs_summary_rows[0] if bs_summary_rows and len(bs_summary_rows) > 0 else []

    def _vl(hdrs, name):
        for i, h in enumerate(hdrs):
            if str(h or "").strip() == name:
                return _gcl_val(i + 1)
        return None

    IS_ACCT_COL = _vl(_is_sum_hdr, "Account Name") or "B"
    IS_AMT_COL  = _vl(_is_sum_hdr, "Amount") or "E"
    IS_DIM_COL  = None

    BS_ACCT_COL = _vl(_bs_sum_hdr, "Account Name") or "B"
    BS_BAL_COL  = _vl(_bs_sum_hdr, "Amount") or "D"

    # Parse dates for Excel DATE() formula
    from datetime import datetime as _dt
    if start_date and end_date:
        sd = _dt.strptime(start_date, "%Y-%m-%d")
        ed = _dt.strptime(end_date, "%Y-%m-%d")
        date_start_formula = f"DATE({sd.year},{sd.month},{sd.day})"
        date_end_formula   = f"DATE({ed.year},{ed.month},{ed.day})"
    else:
        date_start_formula = None
        date_end_formula   = None

    def _tab_ref(tab_name):
        if any(c in tab_name for c in (' ', '-', '/', '&', "'")):
            return f"'{tab_name}'"
        return tab_name

    is_ref = _tab_ref(is_tab_name)
    bs_ref = _tab_ref(bs_tab_name)

    import re as _re_val

    def _bare_lower(name):
        # Strip leading spaces first (P&L/BS rows have indent spaces), then digit prefix
        s = str(name or "").strip()
        s = _re_val.sub(r'^\d[\d.\-]*\s+', '', s).strip()
        return s.lower()

    # Build P&L account → row number lookup
    # Key on FULL account name (stripped of indent spaces only, keeping number prefix)
    # Data rows take priority over Header/Section rows with the same bare name
    pl_row_lookup: dict[str, int] = {}       # full name → row
    pl_bare_lookup: dict[str, int] = {}      # bare name → row (fallback, only if unique data row)
    pl_bare_counts: dict[str, int] = {}      # count bare name occurrences among DATA rows only
    if pl_rows and len(pl_rows) > 1:
        # First pass: count data rows per bare name and build full-name lookup
        # pl_rows data rows have row_type as last element
        num_header_cols = len(pl_rows[0])
        for ri, row in enumerate(pl_rows[1:], 2):
            if not row:
                continue
            full     = str(row[0] or "").strip()
            bare     = _bare_lower(full)
            row_type = str(row[num_header_cols]) if len(row) > num_header_cols else ""
            is_data  = row_type in ("Data", "Section")
            # Exact full-name lookup: prefer data rows, only overwrite with non-data if nothing else
            if full.lower() not in pl_row_lookup or is_data:
                pl_row_lookup[full.lower()] = ri
            # Count data rows per bare name
            if is_data:
                pl_bare_counts[bare] = pl_bare_counts.get(bare, 0) + 1
        # Second pass: build bare lookup only for unique data-row bare names
        for ri, row in enumerate(pl_rows[1:], 2):
            if not row:
                continue
            full     = str(row[0] or "").strip()
            bare     = _bare_lower(full)
            row_type = str(row[num_header_cols]) if len(row) > num_header_cols else ""
            is_data  = row_type in ("Data", "Section")
            if is_data and pl_bare_counts.get(bare, 0) == 1:
                pl_bare_lookup[bare] = ri

    # Build BS account → row number lookup (same approach)
    bs_row_lookup: dict[str, int] = {}
    bs_bare_lookup: dict[str, int] = {}
    bs_bare_counts: dict[str, int] = {}
    bs_num_cols = 0  # number of month columns in Balance Sheet tab
    if bs_report_rows and len(bs_report_rows) > 1:
        bs_num_cols = len(bs_report_rows[0]) - 1 if bs_report_rows[0] else 0
        num_bs_cols = bs_num_cols
        from openpyxl.utils import get_column_letter as _gcl_v
        bs_last_col = _gcl_v(num_bs_cols + 1) if num_bs_cols > 0 else "B"
        for ri, row in enumerate(bs_report_rows[1:], 2):
            if row:
                full = str(row[0] or "").strip()
                bare = _bare_lower(full)
                bs_row_lookup[full.lower()] = ri
                bs_bare_counts[bare] = bs_bare_counts.get(bare, 0) + 1
        for ri, row in enumerate(bs_report_rows[1:], 2):
            if row:
                full = str(row[0] or "").strip()
                bare = _bare_lower(full)
                if bs_bare_counts.get(bare, 0) == 1:
                    bs_bare_lookup[bare] = ri
    else:
        bs_last_col = "B"

    # P&L last column letter (sum all months = total period activity)
    pl_num_cols = len(pl_rows[0]) - 1 if pl_rows and pl_rows[0] else 0
    from openpyxl.utils import get_column_letter as _gcl_v2
    pl_last_col = _gcl_v2(pl_num_cols + 1) if pl_num_cols > 0 else "B"

    headers = ["Account", "Statement", "GL Value (Live)", "QBO Report Value", "Difference", "Status", "Notes"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = BOLD_WHITE
        c.fill = HEADER_BG
        c.alignment = Alignment(horizontal="center")

    SUMMARY_START = 2
    DETAIL_START  = 8
    ws.cell(row=DETAIL_START - 1, column=1, value="ACCOUNT DETAIL").font = BOLD

    current_row = DETAIL_START

    # IS section header
    c = ws.cell(row=current_row, column=1, value="INCOME STATEMENT ACCOUNTS")
    c.font = BOLD
    c.fill = SECTION_BG
    for ci in range(1, 8):
        ws.cell(row=current_row, column=ci).fill = SECTION_BG
    current_row += 1

    def _bare(name):
        return _re.sub(r'^\d[\d.\-]*\s+', '', str(name or "")).strip()

    # Use full account name as-is — each account is distinct (e.g. "65110 Software & Apps"
    # and "Software & apps" are different accounts and must not be merged)
    seen_full = set()
    is_accounts = []
    for acct, val in sorted(qbo_is.items()):
        if acct.lower() not in seen_full:
            seen_full.add(acct.lower())
            amount   = val["amount"] if isinstance(val, dict) else val
            acct_num = val.get("acct_num", "") if isinstance(val, dict) else ""
            is_accounts.append((acct, amount, acct_num))

    for acct, qbo_val, acct_num in is_accounts:
        if abs(qbo_val) < 0.01:
            continue
        rn = current_row
        ws.cell(row=rn, column=1, value=acct)
        ws.cell(row=rn, column=2, value="IS")
        # GL Value: SUMIFS matching account name in col A of Validation against IS GL Detail
        # Always use A{rn} cell reference — never hardcode account names or numbers
        if date_start_formula and date_end_formula:
            _is_formula = (
                f'=SUMIFS({is_ref}!{IS_AMT_COL}:{IS_AMT_COL},'
                f'{is_ref}!{IS_ACCT_COL}:{IS_ACCT_COL},A{rn},'
                f'{is_ref}!A:A,">="&{date_start_formula},'
                f'{is_ref}!A:A,"<="&{date_end_formula})')
        else:
            _is_formula = (
                f'=SUMIF({is_ref}!{IS_ACCT_COL}:{IS_ACCT_COL},'
                f'A{rn},{is_ref}!{IS_AMT_COL}:{IS_AMT_COL})')
        ws.cell(row=rn, column=3, value=_is_formula).number_format = NUM_FMT
        # QBO Report Value: XLOOKUP on account name against P&L Total column
        # P&L account names are stored without leading spaces (indent via cell formatting)
        qbo_formula = f"=IFERROR(_xlfn.XLOOKUP(A{rn},'P&L'!A:A,'P&L'!{pl_last_col}:{pl_last_col},0),0)"
        c = ws.cell(row=rn, column=4, value=qbo_formula)
        c.number_format = NUM_FMT
        ws.cell(row=rn, column=5, value=f"=C{rn}-D{rn}").number_format = NUM_FMT
        ws.cell(row=rn, column=6,
                value=f'=IF(ABS(E{rn})<{TOLERANCE},"MATCH",IF(C{rn}=0,"MISSING","DIFF"))')
        ws.cell(row=rn, column=7, value="")
        current_row += 1

    current_row += 1

    # BS section header
    c = ws.cell(row=current_row, column=1, value="BALANCE SHEET ACCOUNTS")
    c.font = BOLD
    c.fill = SECTION_BG
    for ci in range(1, 8):
        ws.cell(row=current_row, column=ci).fill = SECTION_BG
    current_row += 1

    seen_full_bs = set()
    bs_accounts = []
    for acct, val in sorted(qbo_bs.items()):
        if acct.lower() not in seen_full_bs:
            seen_full_bs.add(acct.lower())
            amount   = val["amount"] if isinstance(val, dict) else val
            acct_num = val.get("acct_num", "") if isinstance(val, dict) else ""
            bs_accounts.append((acct, amount, acct_num))

    for acct, qbo_val, acct_num in bs_accounts:
        if abs(qbo_val) < 0.01:
            continue
        rn = current_row
        ws.cell(row=rn, column=1, value=acct)
        ws.cell(row=rn, column=2, value="BS")
        # BS Balances: A=Account, B=Account Type, C=Subtype, D=Account Group, E=Month, F=Ending Balance
        # Exact name match — no wildcard
        if date_end_formula:
            _bs_formula = (
                f"=IFERROR(SUMIFS('BS Balances'!F:F,"
                f"'BS Balances'!A:A,A{rn},"
                f"'BS Balances'!E:E,{date_end_formula}),0)")
        else:
            _bs_formula = (
                f"=IFERROR(SUMIF('BS Balances'!A:A,A{rn},'BS Balances'!F:F),0)")
        ws.cell(row=rn, column=3, value=_bs_formula).number_format = NUM_FMT
        # QBO Report Value: XLOOKUP on account name against Balance Sheet last month column
        qbo_formula = f"=IFERROR(_xlfn.XLOOKUP(A{rn},'Balance Sheet'!A:A,'Balance Sheet'!{bs_last_col}:{bs_last_col},0),0)"
        c = ws.cell(row=rn, column=4, value=qbo_formula)
        c.number_format = NUM_FMT
        ws.cell(row=rn, column=5, value=f"=C{rn}-D{rn}").number_format = NUM_FMT
        ws.cell(row=rn, column=6,
                value=f'=IF(ABS(E{rn})<{TOLERANCE},"MATCH",IF(AND(C{rn}=0,B{rn}="BS"),"NO ACTIVITY","DIFF"))')
        ws.cell(row=rn, column=7, value="")
        current_row += 1

    last_data_row = current_row - 1

    # Summary formulas (rows 2-6)
    summary_data = [
        ("Overall Result",
         f'=IF(COUNTIF(F{DETAIL_START}:F{last_data_row},"DIFF")+COUNTIF(F{DETAIL_START}:F{last_data_row},"MISSING")=0,"\u2713 PASS","\u2717 FAIL")', ""),
        ("Total Accounts Checked",
         f'=COUNTA(A{DETAIL_START}:A{last_data_row})-COUNTIF(A{DETAIL_START}:A{last_data_row},"INCOME STATEMENT ACCOUNTS")-COUNTIF(A{DETAIL_START}:A{last_data_row},"BALANCE SHEET ACCOUNTS")', ""),
        ("Matched", f'=COUNTIF(F{DETAIL_START}:F{last_data_row},"MATCH")', ""),
        ("Differences Found", f'=COUNTIF(F{DETAIL_START}:F{last_data_row},"DIFF")', "Red rows below"),
        ("Missing from GL", f'=COUNTIF(F{DETAIL_START}:F{last_data_row},"MISSING")', "Yellow rows below"),
    ]
    for i, (label, formula, note) in enumerate(summary_data):
        r = SUMMARY_START + i
        ws.cell(row=r, column=1, value=label).font = BOLD
        ws.cell(row=r, column=3, value=formula)
        if note:
            ws.cell(row=r, column=7, value=note)

    # Conditional formatting
    ws.conditional_formatting.add(
        f"C{SUMMARY_START}",
        CellIsRule(operator="equal", formula=['"\u2713 PASS"'],
                   fill=PatternFill("solid", fgColor="C6EFCE"), font=Font(bold=True)))
    ws.conditional_formatting.add(
        f"C{SUMMARY_START}",
        CellIsRule(operator="equal", formula=['"\u2717 FAIL"'],
                   fill=PatternFill("solid", fgColor="FFC7CE"), font=Font(bold=True)))

    detail_range = f"A{DETAIL_START}:G{last_data_row}"
    # FormulaRule: $F locks column, row is relative — Excel extends down each row
    ws.conditional_formatting.add(detail_range,
        FormulaRule(formula=[f'$F{DETAIL_START}="MATCH"'],
                    fill=GREEN, font=Font(color="276221")))
    ws.conditional_formatting.add(detail_range,
        FormulaRule(formula=[f'$F{DETAIL_START}="DIFF"'],
                    fill=RED, font=Font(color="9C0006")))
    ws.conditional_formatting.add(detail_range,
        FormulaRule(formula=[f'$F{DETAIL_START}="MISSING"'],
                    fill=YELLOW, font=Font(color="9C5700")))
    GREY = PatternFill("solid", fgColor="F2F2F2")
    ws.conditional_formatting.add(detail_range,
        FormulaRule(formula=[f'$F{DETAIL_START}="NO ACTIVITY"'], fill=GREY))

    widths = [48, 12, 18, 20, 14, 12, 35]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A2"
    progress_fn(f"  Validation tab written: {len(is_accounts)} IS accounts, {len(bs_accounts)} BS accounts")


def _write_report_sheet(
    wb: openpyxl.Workbook,
    tab_name: str,
    rows: list[list],
    report_type: str = "P&L",
    validation_rows: list[list] | None = None,
):
    """Write a monthly P&L or Balance Sheet tab with formatting."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if not rows or len(rows) < 2:
        ws = wb.create_sheet(tab_name)
        ws.sheet_view.showGridLines = False
        ws.cell(row=1, column=1, value="No data")
        return

    ws = wb.create_sheet(tab_name)
    ws.sheet_view.showGridLines = False

    # Header row
    header = rows[0]
    for ci, val in enumerate(header, 1):
        c = ws.cell(row=1, column=ci, value=val)
        c.font      = HDR_FONT
        c.fill      = HDR_FILL
        c.alignment = Alignment(horizontal="center" if ci > 1 else "left")

    GRANDTOTAL_KEYWORDS = {
        "net income", "gross profit", "net operating income",
        "net other income", "total equity", "total liabilities and equity",
    }
    SUBTOTAL_PREFIXES = ("total ", "net ")

    for ri, row in enumerate(rows[1:], 2):
        try:
            if not row:
                continue
            num_data_cols = len(rows[0])
            if len(row) > num_data_cols:
                row = row[:num_data_cols]

            label     = str(row[0] or "")
            bare      = label.strip().lower()
            indent_lv = (len(label) - len(label.lstrip())) // 2

            is_grandtotal = bare in GRANDTOTAL_KEYWORDS
            is_subtotal   = (any(bare.startswith(p) for p in SUBTOTAL_PREFIXES)
                             and not is_grandtotal)
            is_header     = (indent_lv == 0
                             and not is_grandtotal
                             and not is_subtotal
                             and not any(c.isdigit() for c in bare[:5]))

            if is_header:
                c = ws.cell(row=ri, column=1, value=label.strip())
                c.font = BOLD_FONT
                c.fill = GRAY_FILL
                for ci in range(2, len(row) + 1):
                    c = ws.cell(row=ri, column=ci, value=None)
                    c.font = PLAIN_FONT
                    c.fill = GRAY_FILL
                continue

            for ci in range(1, len(row) + 1):
                val = row[ci - 1] if (ci - 1) < len(row) else None
                if val is not None and not isinstance(val, (str, int, float, date)):
                    val = str(val)
                if ci == 1 and isinstance(val, str):
                    val = val.strip()

                c = ws.cell(row=ri, column=ci, value=val)
                c.font = BOLD_FONT if (is_grandtotal or is_subtotal or is_header) else PLAIN_FONT

                if ci == 1:
                    c.alignment = Alignment(horizontal="left", indent=indent_lv)
                elif isinstance(val, (int, float)):
                    c.number_format = _ACCT_FMT
                    c.alignment     = Alignment(horizontal="right")

                if is_grandtotal:
                    c.fill   = NO_FILL
                    c.border = _Bdr(top=THIN, bottom=DOUBLE)
                elif is_subtotal:
                    c.fill   = NO_FILL
                    c.border = _Bdr(top=THIN)
                elif is_header:
                    c.fill = GRAY_FILL
                else:
                    c.fill = NO_FILL

        except Exception as e:
            print(f"  WARNING: _write_report_sheet skipped row {ri}: {e}")
            continue

    # Cross-check validation rows at the bottom
    if validation_rows:
        VAL_BG   = PatternFill("solid", fgColor="FFFDE7")
        VAL_BOLD = _font(bold=True)

        separator_row  = len(rows) + 2
        ws.cell(row=separator_row, column=1, value="")
        val_header_row = separator_row + 1
        ws.cell(row=val_header_row, column=1, value="\u2500\u2500 GL CROSS-CHECK \u2500\u2500").font = VAL_BOLD

        for vi, vrow in enumerate(validation_rows):
            vri = val_header_row + 1 + vi
            is_diff_row = str(vrow[0] if vrow else "").startswith("Difference")
            for ci, val in enumerate(vrow, 1):
                c = ws.cell(row=vri, column=ci, value=val)
                c.fill = VAL_BG
                if ci == 1:
                    c.font = _font(bold=True) if vi == 0 else PLAIN_FONT
                if ci > 1 and isinstance(val, str) and val.startswith("="):
                    c.number_format = "#,##0.00"
            # Add conditional formatting to Difference rows — red if non-zero
            if is_diff_row and len(vrow) > 1:
                from openpyxl.formatting.rule import CellIsRule
                diff_range = f"B{vri}:{get_column_letter(len(vrow))}{vri}"
                ws.conditional_formatting.add(
                    diff_range,
                    CellIsRule(
                        operator="notEqual",
                        formula=["0"],
                        font=Font(bold=True, color="C00000"),
                        fill=PatternFill("solid", fgColor="FFE0E0"),
                    )
                )

    ws.column_dimensions["A"].width = 42
    for ci in range(2, len(rows[0]) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    ws.freeze_panes = "B2"


def _write_dimension_sheet(
    wb: openpyxl.Workbook,
    tab_name: str,
    rows: list[list],
    dim_label: str = "Class",
):
    """Write an IS by Class or IS by Location tab."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    if not rows or len(rows) < 2:
        ws = wb.create_sheet(tab_name)
        ws.sheet_view.showGridLines = False
        ws.cell(row=1, column=1, value="No data")
        return

    ws = wb.create_sheet(tab_name)
    ws.sheet_view.showGridLines = False

    header = rows[0]
    for ci, val in enumerate(header, 1):
        c           = ws.cell(row=1, column=ci, value=val)
        c.font      = HDR_FONT
        c.fill      = HDR_FILL
        c.alignment = _AL(horizontal="center" if ci > 4 else "left", vertical="center")

    for ri, row in enumerate(rows[1:], 2):
        if not row:
            continue
        for ci, val in enumerate(row, 1):
            if val is not None and not isinstance(val, (str, int, float, date)):
                val = str(val)
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = PLAIN_FONT
            if ci > 4 and isinstance(val, (int, float)):
                c.number_format = _ACCT_FMT
                c.alignment     = _AL(horizontal="right")

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 25
    for ci in range(5, len(rows[0]) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    ws.column_dimensions["C"].hidden = True
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(rows[0]))}{len(rows)}"


# ── Main entry point ─────────────────────────────────────────────────────────

def _build_is_gl_summary(is_gl_rows: list[list]) -> list[list]:
    """Aggregate IS GL Detail rows by Account + Month + Dimension."""
    if not is_gl_rows or len(is_gl_rows) < 2:
        return is_gl_rows
    from collections import defaultdict
    header = is_gl_rows[0]
    def ci(name):
        try: return header.index(name)
        except ValueError: return -1
    i_an = ci("Account Name"); i_mo = ci("Month")
    i_amt = ci("Amount"); i_at = ci("Account Type")
    i_dim = -1; dim_name = None
    for dn in ("Class", "Location", "Department"):
        i_dim = ci(dn)
        if i_dim >= 0: dim_name = dn; break
    mgc = [(i, h) for i, h in enumerate(header) if str(h or "").endswith(" - Account Group")]
    msc = [(i, h) for i, h in enumerate(header) if str(h or "").endswith(" - Statement Section")]
    agg = defaultdict(float); meta = {}
    for row in is_gl_rows[1:]:
        an = row[i_an] if i_an >= 0 else ""
        mo = row[i_mo] if i_mo >= 0 else ""; dv = row[i_dim] if i_dim >= 0 else "Total"
        am = float(row[i_amt] or 0) if i_amt >= 0 else 0.0
        key = (an, mo, dv); agg[key] += am
        if key not in meta:
            meta[key] = {
                "at": row[i_at] if i_at >= 0 else "",
                "mi": {i: (row[i] if i < len(row) else "") for i, _ in mgc + msc},
            }
    sh = ["Month", "Account Name", "Account Type"]
    if dim_name: sh.append(dim_name)
    sh.append("Amount")
    for _, h in mgc: sh.append(h)
    for _, h in msc: sh.append(h)
    out = [sh]
    for key, amt in sorted(agg.items(), key=lambda x: (str(x[0][1]), str(x[0][0]))):
        an, mo, dv = key; m = meta[key]
        r = [mo, an, m["at"]]
        if dim_name: r.append(dv)
        r.append(amt)
        for i, _ in mgc: r.append(m["mi"].get(i, ""))
        for i, _ in msc: r.append(m["mi"].get(i, ""))
        out.append(r)
    return out


def _build_bs_gl_summary(bs_gl_rows: list[list]) -> list[list]:
    """Aggregate BS GL Detail rows by Account + Month."""
    if not bs_gl_rows or len(bs_gl_rows) < 2:
        return bs_gl_rows
    from collections import defaultdict
    header = bs_gl_rows[0]
    def ci(name):
        try: return header.index(name)
        except ValueError: return -1
    i_an = ci("Account Name"); i_mo = ci("Month")
    i_amt = ci("Amount"); i_at = ci("Account Type")
    mgc = [(i, h) for i, h in enumerate(header) if str(h or "").endswith(" - Account Group")]
    msc = [(i, h) for i, h in enumerate(header) if str(h or "").endswith(" - Statement Section")]
    agg = defaultdict(float); meta = {}
    for row in bs_gl_rows[1:]:
        an = row[i_an] if i_an >= 0 else ""
        mo = row[i_mo] if i_mo >= 0 else ""
        am = float(row[i_amt] or 0) if i_amt >= 0 else 0.0
        key = (an, mo); agg[key] += am
        if key not in meta:
            meta[key] = {
                "at": row[i_at] if i_at >= 0 else "",
                "mi": {i: (row[i] if i < len(row) else "") for i, _ in mgc + msc},
            }
    sh = ["Month", "Account Name", "Account Type", "Amount"]
    for _, h in mgc: sh.append(h)
    for _, h in msc: sh.append(h)
    out = [sh]
    for key, amt in sorted(agg.items(), key=lambda x: (str(x[0][1]), str(x[0][0]))):
        an, mo = key; m = meta[key]
        r = [mo, an, m["at"], amt]
        for i, _ in mgc: r.append(m["mi"].get(i, ""))
        for i, _ in msc: r.append(m["mi"].get(i, ""))
        out.append(r)
    return out


def generate_lite(
    alias: str,
    start_date: str,
    end_date: str,
    output_mode: str = "new",
    output_folder: str = "",
    file_name: str = "",
    existing_file: str = "",
    progress_fn: Callable = print,
    pct_fn: Callable[[int], None] | None = None,
    cancel_fn: Callable[[], bool] | None = None,
    dimension: str = "class",
    include_gl_detail: bool = False,
    include_ar_aging: bool = False,
    include_ap_aging: bool = False,
) -> dict:
    """
    Run Acorn Lite extraction — three pulls to one Excel file.

    output_mode:
      "new"      — create a new file in output_folder (with optional file_name)
      "existing" — write tabs into an existing .xlsx file

    Returns {"path": str} with the output file path.
    """
    alias = alias.upper()
    progress_fn(f"\n  Acorn Lite — {alias}")
    progress_fn(f"  Period: {start_date} to {end_date}\n")

    _check_cancel(cancel_fn)
    if pct_fn: pct_fn(5)

    progress_fn("  Fetching Chart of Accounts...")
    coa_lookup = _build_coa_lookup(alias, progress_fn)
    progress_fn(f"  COA: {len(coa_lookup)} accounts")

    _check_cancel(cancel_fn)
    if pct_fn: pct_fn(10)

    progress_fn("\n  Fetching GL...")
    gl_df = _fetch_gl(alias, start_date, end_date, progress_fn)

    _check_cancel(cancel_fn)
    if pct_fn: pct_fn(40)

    progress_fn("\n  Building IS GL Detail...")
    is_rows = _prepare_is_gl_rows(gl_df, coa_lookup, progress_fn, dimension=dimension)
    n_is = len(is_rows) - 1 if is_rows else 0
    progress_fn(f"[progress] IS GL: {n_is} rows")

    _check_cancel(cancel_fn)
    if pct_fn: pct_fn(50)

    progress_fn("\n  Building BS GL Detail...")
    bs_rows = _prepare_bs_gl_rows(gl_df, coa_lookup, progress_fn)
    n_bs = len(bs_rows) - 1 if bs_rows else 0
    progress_fn(f"[progress] BS GL: {n_bs} rows")

    _check_cancel(cancel_fn)
    if pct_fn: pct_fn(58)

    progress_fn("\n  Fetching BS Balances...")
    bal_rows = _fetch_bs_balances(alias, start_date, end_date, progress_fn, coa_lookup=coa_lookup)
    n_bal = len(bal_rows) - 1 if bal_rows else 0
    progress_fn(f"[progress] BS Balances: {n_bal} rows")
    if pct_fn: pct_fn(65)

    _check_cancel(cancel_fn)
    progress_fn("\n  Fetching monthly P&L and Balance Sheet...")
    pl_report_rows, bs_report_rows = _fetch_monthly_reports(
        alias, start_date, end_date, progress_fn
    )
    n_pl = len(pl_report_rows) - 1 if pl_report_rows else 0
    n_bs_report = len(bs_report_rows) - 1 if bs_report_rows else 0
    progress_fn(f"[progress] P&L: {n_pl} rows, Balance Sheet: {n_bs_report} rows")
    if pct_fn: pct_fn(80)

    # Resolve output path
    if output_mode == "existing" and existing_file:
        save_path = Path(existing_file)
        if save_path.exists():
            progress_fn(f"\n  Writing to existing file: {save_path.name}...")
            wb = openpyxl.load_workbook(save_path)
            wb.calculation.calcMode = "auto"
            wb.calculation.fullCalcOnLoad = True
            # Remove old tabs if they exist, then recreate
            for tab_name in ("IS GL Detail", "BS GL Detail", "BS Balances", "AR Aging", "AP Aging", "P&L", "Balance Sheet", "Validation"):
                if tab_name in wb.sheetnames:
                    del wb[tab_name]
        else:
            progress_fn(f"\n  File not found, creating: {save_path.name}...")
            wb = openpyxl.Workbook()
            wb.calculation.calcMode = "auto"
            wb.calculation.fullCalcOnLoad = True
            wb.remove(wb.active)
    else:
        out_dir = Path(output_folder)
        out_dir.mkdir(parents=True, exist_ok=True)
        if file_name:
            # Ensure .xlsx extension
            if not file_name.lower().endswith(".xlsx"):
                file_name += ".xlsx"
            save_path = out_dir / file_name
        else:
            start_ym = start_date[:7]
            end_ym = end_date[:7]
            save_path = out_dir / f"{alias}_GL_{start_ym}_{end_ym}.xlsx"
        progress_fn(f"\n  Writing {save_path.name}...")
        wb = openpyxl.Workbook()
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.remove(wb.active)

    # Set default font for entire workbook
    if "Normal" in wb.style_names:
        wb._named_styles["Normal"].font = _Font(name=_ARIAL, size=_FONT_SZ)

    # Build summary tabs
    is_summary_rows = _build_is_gl_summary(is_rows)
    bs_summary_rows = _build_bs_gl_summary(bs_rows)

    # Write tabs in order — summary always, detail only if requested
    _write_sheet(wb, "IS GL Summary", is_summary_rows)
    _write_sheet(wb, "BS GL Summary", bs_summary_rows)
    if include_gl_detail:
        _write_sheet(wb, "IS GL Detail", is_rows)
        _write_sheet(wb, "BS GL Detail", bs_rows)
    _write_sheet(wb, "BS Balances", bal_rows)

    # ── AR / AP Aging tabs (optional) ─────────────────────────────────────
    if include_ar_aging:
        try:
            progress_fn("\n  Fetching AR Aging...")
            ar_rows = _fetch_ar_aging(alias, end_date, progress_fn)
            if ar_rows and len(ar_rows) > 1:
                _write_aging_sheet(wb, "AR Aging", ar_rows, end_date)
        except Exception as _ar_err:
            progress_fn(f"  WARNING: AR Aging failed — {_ar_err}")

    if include_ap_aging:
        try:
            progress_fn("\n  Fetching AP Aging...")
            ap_rows = _fetch_ap_aging(alias, end_date, progress_fn)
            if ap_rows and len(ap_rows) > 1:
                _write_aging_sheet(wb, "AP Aging", ap_rows, end_date)
        except Exception as _ap_err:
            progress_fn(f"  WARNING: AP Aging failed — {_ap_err}")

    # Build P&L cross-check: Net Income from P&L report vs IS GL Summary SUMIFS
    # Dynamic column lookup — resolve column letters from summary headers
    def _col_letter(headers, col_name):
        from openpyxl.utils import get_column_letter as _gcl
        for i, h in enumerate(headers):
            if str(h or "").strip() == col_name:
                return _gcl(i + 1)
        return None

    is_sum_hdr = is_summary_rows[0] if is_summary_rows else []
    IS_AMT = _col_letter(is_sum_hdr, "Amount")
    IS_MON = _col_letter(is_sum_hdr, "Month")
    IS_TYP = _col_letter(is_sum_hdr, "Account Type")

    pl_validation_rows = []
    if pl_report_rows and len(pl_report_rows) > 1 and IS_AMT and IS_MON and IS_TYP:
        month_labels_val = pl_report_rows[0][1:]
        num_months_val   = len(month_labels_val)

        from openpyxl.utils import get_column_letter as _gcl_pl

        net_income_row_idx = None
        for ri, row in enumerate(pl_report_rows):
            label = str(row[0]).strip().lower()
            if label in ("net income", "net earnings"):
                net_income_row_idx = ri + 1
                break

        if net_income_row_idx:
            pl_ni_row = ["Net Income \u2014 P&L Report"] + [
                f"={_gcl_pl(ci+2)}{net_income_row_idx}"
                for ci in range(num_months_val)
            ]

            gl_ni_row = ["Net Income \u2014 IS GL Summary"]
            from datetime import datetime as _dt_val
            import calendar as _cal_val
            for ml in month_labels_val:
                try:
                    month_dt = _dt_val.strptime(ml, "%b %Y")
                    me_day = _cal_val.monthrange(month_dt.year, month_dt.month)[1]
                    me = f"DATE({month_dt.year},{month_dt.month},{me_day})"
                    income_formula = (
                        f"SUMIFS('IS GL Summary'!${IS_AMT}:${IS_AMT},"
                        f"'IS GL Summary'!${IS_MON}:${IS_MON},{me},"
                        f"'IS GL Summary'!${IS_TYP}:${IS_TYP},\"Income\")"
                        f"+SUMIFS('IS GL Summary'!${IS_AMT}:${IS_AMT},"
                        f"'IS GL Summary'!${IS_MON}:${IS_MON},{me},"
                        f"'IS GL Summary'!${IS_TYP}:${IS_TYP},\"Other Income\")"
                    )
                    expense_formula = (
                        f"SUMIFS('IS GL Summary'!${IS_AMT}:${IS_AMT},"
                        f"'IS GL Summary'!${IS_MON}:${IS_MON},{me},"
                        f"'IS GL Summary'!${IS_TYP}:${IS_TYP},\"Expense\")"
                        f"+SUMIFS('IS GL Summary'!${IS_AMT}:${IS_AMT},"
                        f"'IS GL Summary'!${IS_MON}:${IS_MON},{me},"
                        f"'IS GL Summary'!${IS_TYP}:${IS_TYP},\"Cost of Goods Sold\")"
                        f"+SUMIFS('IS GL Summary'!${IS_AMT}:${IS_AMT},"
                        f"'IS GL Summary'!${IS_MON}:${IS_MON},{me},"
                        f"'IS GL Summary'!${IS_TYP}:${IS_TYP},\"Other Expense\")"
                    )
                    formula = f"=({income_formula})-({expense_formula})"
                except Exception:
                    formula = 0.0
                gl_ni_row.append(formula)

            R1_pl = len(pl_report_rows) + 4
            R2_pl = len(pl_report_rows) + 5
            diff_row_pl = ["Difference (should be zero)"] + [
                f"={_gcl_pl(ci+2)}{R1_pl}-{_gcl_pl(ci+2)}{R2_pl}"
                for ci in range(num_months_val)
            ]

            pl_validation_rows = [pl_ni_row, gl_ni_row, diff_row_pl]

    # Build BS cross-check: Total Assets/Liabilities from BS Report vs BS Balances
    bs_validation_rows = []
    if bs_report_rows and len(bs_report_rows) > 1:
        bs_month_labels_val = bs_report_rows[0][1:]
        num_bs_months_val   = len(bs_month_labels_val)

        from openpyxl.utils import get_column_letter as _gcl_bs
        from datetime import datetime as _dt_bs
        import calendar as _cal_bs

        # Parse BS column labels back to month-end dates
        # Labels are like "Aug 31, 2025"
        bs_month_dates = []
        for lbl in bs_month_labels_val:
            lbl = str(lbl).strip()
            parsed = None
            for fmt in ("%b %d, %Y", "%b %Y", "%Y-%m-%d", "%Y-%m"):
                try:
                    d = _dt_bs.strptime(lbl, fmt)
                    last_day = _cal_bs.monthrange(d.year, d.month)[1]
                    parsed = date(d.year, d.month, last_day)
                    break
                except (ValueError, TypeError):
                    continue
            bs_month_dates.append(parsed)

        # Find Total Assets, Total Liabilities, Total Equity rows
        total_assets_idx = None
        total_liab_idx   = None
        total_equity_idx = None
        for ri, row in enumerate(bs_report_rows):
            label = str(row[0]).strip().lower()
            if "total assets" in label and total_assets_idx is None:
                total_assets_idx = ri + 1
            if "total liabilities" in label and "equity" not in label and total_liab_idx is None:
                total_liab_idx = ri + 1
            if "total equity" in label and total_equity_idx is None:
                total_equity_idx = ri + 1

        ASSET_TYPES = ["Bank", "Accounts Receivable", "Other Current Asset",
                       "Fixed Asset", "Other Asset"]
        LIAB_TYPES  = ["Accounts Payable", "Credit Card", "Other Current Liability",
                       "Long Term Liability"]

        # Total Assets rows
        ta_report_row = ["Total Assets — Balance Sheet Report"] + (
            [f"={_gcl_bs(ci+2)}{total_assets_idx}" for ci in range(num_bs_months_val)]
            if total_assets_idx else [""] * num_bs_months_val
        )
        ta_gl_row = ["Total Assets — BS Balances"]
        for me_dt in bs_month_dates:
            if me_dt:
                me_formula = f"DATE({me_dt.year},{me_dt.month},{me_dt.day})"
                ta_gl_row.append(
                    f"=SUMIFS('BS Balances'!F:F,"
                    f"'BS Balances'!E:E,{me_formula},"
                    f"'BS Balances'!D:D,\"Asset\")"
                )
            else:
                ta_gl_row.append("")

        # Total Liabilities rows
        tl_report_row = ["Total Liabilities — Balance Sheet Report"] + (
            [f"={_gcl_bs(ci+2)}{total_liab_idx}" for ci in range(num_bs_months_val)]
            if total_liab_idx else [""] * num_bs_months_val
        )
        tl_gl_row = ["Total Liabilities — BS Balances"]
        for me_dt in bs_month_dates:
            if me_dt:
                me_formula = f"DATE({me_dt.year},{me_dt.month},{me_dt.day})"
                tl_gl_row.append(
                    f"=SUMIFS('BS Balances'!F:F,"
                    f"'BS Balances'!E:E,{me_formula},"
                    f"'BS Balances'!D:D,\"Liability\")"
                )
            else:
                tl_gl_row.append("")

        # Total Equity rows
        te_report_row = ["Total Equity — Balance Sheet Report"] + (
            [f"={_gcl_bs(ci+2)}{total_equity_idx}" for ci in range(num_bs_months_val)]
            if total_equity_idx else [""] * num_bs_months_val
        )
        te_gl_row = ["Total Equity — BS Balances"]
        for me_dt in bs_month_dates:
            if me_dt:
                me_formula = f"DATE({me_dt.year},{me_dt.month},{me_dt.day})"
                te_gl_row.append(
                    f"=SUMIFS('BS Balances'!F:F,"
                    f"'BS Balances'!E:E,{me_formula},"
                    f"'BS Balances'!D:D,\"Equity\")"
                )
            else:
                te_gl_row.append("")

        # Diff rows — reference actual Excel row numbers
        # vi=0: ta_report, vi=1: ta_gl, vi=2: ta_diff
        # vi=3: spacer,    vi=4: tl_report, vi=5: tl_gl, vi=6: tl_diff
        # vi=7: spacer,    vi=8: te_report, vi=9: te_gl, vi=10: te_diff
        BASE = len(bs_report_rows) + 4
        R1 = BASE + 0;  R2 = BASE + 1   # assets
        R3 = BASE + 4;  R4 = BASE + 5   # liabilities
        R5 = BASE + 8;  R6 = BASE + 9   # equity
        ta_diff_row = ["Difference — Assets (should be zero)"] + [
            f"={_gcl_bs(ci+2)}{R1}-{_gcl_bs(ci+2)}{R2}"
            for ci in range(num_bs_months_val)
        ]
        tl_diff_row = ["Difference — Liabilities (should be zero)"] + [
            f"={_gcl_bs(ci+2)}{R3}-{_gcl_bs(ci+2)}{R4}"
            for ci in range(num_bs_months_val)
        ]
        te_diff_row = ["Difference — Equity (should be zero)"] + [
            f"={_gcl_bs(ci+2)}{R5}-{_gcl_bs(ci+2)}{R6}"
            for ci in range(num_bs_months_val)
        ]

        bs_validation_rows = [
            ta_report_row, ta_gl_row, ta_diff_row,
            [""] + [""] * num_bs_months_val,
            tl_report_row, tl_gl_row, tl_diff_row,
            [""] + [""] * num_bs_months_val,
            te_report_row, te_gl_row, te_diff_row,
        ]

    _write_report_sheet(wb, "P&L", pl_report_rows, report_type="P&L",
                        validation_rows=pl_validation_rows)
    _write_report_sheet(wb, "Balance Sheet", bs_report_rows, report_type="Balance Sheet",
                        validation_rows=bs_validation_rows)
    _check_cancel(cancel_fn)
    progress_fn("\n  Fetching QBO report totals for validation...")
    qbo_is, qbo_bs = _fetch_qbo_report_totals(alias, start_date, end_date, progress_fn, coa_lookup=coa_lookup)
    if pct_fn: pct_fn(90)

    progress_fn("\n  Writing Validation tab...")
    _write_validation_sheet(
        wb, qbo_is, qbo_bs,
        start_date=start_date,
        end_date=end_date,
        is_tab_name="IS GL Summary",
        bs_tab_name="BS GL Summary",
        progress_fn=progress_fn,
        dimension=dimension,
        pl_rows=pl_report_rows,
        bs_report_rows=bs_report_rows,
        is_summary_rows=is_summary_rows,
        bs_summary_rows=bs_summary_rows,
    )
    if pct_fn: pct_fn(95)

    try:
        wb.save(save_path)
    except PermissionError:
        # Cache the built workbook so the user can retry without re-pulling data
        _pending_save["wb"]   = wb
        _pending_save["path"] = save_path
        raise PermissionError(
            f"Cannot save to '{save_path.name}' — the file is open in Excel. "
            "Close it and click Try Again."
        )
    # Clear any pending save on success
    _pending_save.clear()
    if pct_fn: pct_fn(100)
    progress_fn(f"  Saved to: {save_path}")

    return {"path": str(save_path)}